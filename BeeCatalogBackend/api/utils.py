import io
import json
import mimetypes
import os
import random
import re
import time
import uuid
import sys
import chardet
from collections import defaultdict
from dotenv import load_dotenv
from datetime import datetime, timedelta
from functools import lru_cache
from typing import Dict, FrozenSet, List, NamedTuple, Optional, Tuple, Union

# Importar sistema de cache inteligente
from .cache_utils import ai_cache, get_or_cache_ai_response, batch_ai_requests

# Importar sistema de memória inteligente de produtos
from .memory_utils import (
    get_cached_content_or_generate,
    save_generated_content_to_memory,
    check_product_in_memory,
    batch_check_products_in_memory,
    extract_product_identifier
)

import requests
from bs4 import BeautifulSoup

import pandas as pd
from django.conf import settings
from django.core.files.storage import FileSystemStorage, default_storage
from django.http import FileResponse, JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.datastructures import MultiValueDictKeyError

import boto3
from botocore.exceptions import ClientError
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from langchain.prompts import PromptTemplate
from langchain_community.document_loaders.csv_loader import CSVLoader
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from langchain_community.vectorstores import FAISS
from langchain_core.output_parsers import JsonOutputParser, StrOutputParser, CommaSeparatedListOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain_core.documents import Document
from pydantic.v1 import BaseModel, Field, constr, conlist

caminho_dotenv = os.path.join(settings.BASE_DIR, ".env")
load_dotenv(dotenv_path=caminho_dotenv)

GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")

AWS_ACCESS_KEY_ID = os.getenv('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.getenv('AWS_SECRET_ACCESS_KEY')
AWS_STORAGE_BUCKET_NAME = os.getenv('AWS_STORAGE_BUCKET_NAME')
AWS_S3_REGION_NAME = os.getenv('AWS_S3_REGION_NAME', 'sa-east-1')

GEMINI_FLASH_INPUT_COST_PER_K_TOKENS = 0.00035
GEMINI_FLASH_OUTPUT_COST_PER_K_TOKENS = 0.00045

s3 = boto3.client(
    's3',
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    region_name=AWS_S3_REGION_NAME
)

THROTTLE_SECONDS = 0.2

DIRETORIO_UPLOAD = os.path.join(settings.BASE_DIR, "uploads")
NOME_TEMPLATE = "AMAZON_TEMPLATE.xlsm"

NOME_CSV_VALORES_VALIDOS = "valores_validos.csv"
NOME_CSV_J_PLANILHA = "J_planilha.csv"
NOME_CSV_J = "J.csv"
NOME_CSV_EXPLORAR_DADOS = "explorar_dados.csv"
NOME_CSV_DEFINICOES_DADOS = "definicoes_dados.csv"

FAISS_INDEX_DIR = os.path.join(settings.BASE_DIR, "faiss_index")
FAISS_INDEX_NAME = "amazon_base"

MAPA_CAMPOS = [
    ("sku", "SKU"),
    ("tipo_marca", "Nome da marca"), 
    ("nome_marca", "Nome da marca"),
    ("preco", "Preço padrão BRL"),
    ("quantidade", "Quantidade (BR)"),
    ("fba_dba", "Código do canal de processamento (BR)"),
    ("id_produto", "ID do produto"),
    ("ncm", "Código NCM"),
    ("peso_pacote", "Peso do pacote"),
    ("peso_produto", "Peso do item"), 
    ("ajuste", "Tipo de ajuste"),
]

MAPA_CLP = { "c_l_a_pacote": [
    "Comprimento do pacote",
    "Largura do pacote",
    "Altura do pacote"
]}

MAPA_CLPR = { "c_l_a_produto": [
    "Comprimento do item",
    "Largura do item",
    "Altura do item"
]}

# CORREÇÃO: Removido MAPA_UNIDADES para evitar preenchimento automático de unidades
# As unidades devem ser preenchidas apenas quando explicitamente necessário
# conforme solicitado pelo usuário
MAPA_UNIDADES = {
    # Mapa vazio - unidades não devem ser preenchidas automaticamente
}

campos_criticos = {
    "Baterias são necessárias?",
    "Quantidade de itens",
    "Cor",
    "Contagem de unidades",
    "Tipo de ID do produto",
    "Caminhos de Navegação Recomendados",
    "País de Origem",
    "Regulamentações de produtos perigosos"
}
HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:112.0) '
        'Gecko/20100101 Firefox/112.0'
    )
}

def format_product_context(product: dict) -> str:
    """Formata os dados do produto em uma string legível para a IA."""
    context_parts = []
    readable_keys = {
        "titulo": "Título do Produto", "sku": "SKU", "tipo_marca": "Tipo de Marca",
        "nome_marca": "Nome da Marca", "preco": "Preço", "fba_dba": "Logística (FBA ou DBA)",
        "id_produto": "ID do Produto (EAN/GTIN/UPC)", "tipo_id_produto": "Tipo de ID do Produto",
        "ncm": "NCM", "quantidade": "Quantidade em Estoque", "peso_pacote": "Peso do Pacote (g)",
        "c_l_a_pacote": "Dimensões do Pacote C x L x A", "peso_produto": "Peso do Produto (g)",
        "c_l_a_produto": "Dimensões do Produto C x L x A", "ajuste": "Produto Ajustável?",
        "tema_variacao_pai": "Tema de Variação Principal"
    }
    for key, value in product.items():
        if value and key in readable_keys:
            context_parts.append(f"- {readable_keys[key]}: {value}")

    variations = product.get('variacoes', [])
    if variations:
        context_parts.append("\n- Variações do Produto:")
        for i, var in enumerate(variations):
            var_details = []
            readable_var_keys = {
                "sku": "SKU da Variação", "tipo": "Tipo de Variação", "cor": "Nome da Cor",
                "cla": "Dimensões (C x L x A)", "peso": "Peso (g)", "imagem": "URL da Imagem"
            }
            for v_key, v_value in var.items():
                if v_value and v_key in readable_var_keys:
                    var_details.append(f"  - {readable_var_keys[v_key]}: {v_value}")
            if var_details:
                context_parts.append(f"  Variação {i+1}:\n" + "\n".join(var_details))
    
    return "\n".join(context_parts)

def scrape_mercadolivre_images(url: str) -> list[str]:
    if "mercadolivre.com.br" not in url and "mercadolivre.com" not in url:
        raise ValueError("A URL fornecida não parece ser do Mercado Livre.")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as e:
        raise ValueError(f"Erro ao acessar a URL: {e}")
    soup = BeautifulSoup(resp.text, 'html.parser')
    gallery = soup.find('div', class_='ui-pdp-gallery')
    images: list[str] = []
    if gallery:
        for img in gallery.find_all('img'):
            src = img.get('data-src') or img.get('src')
            if not src: continue
            high_res_src = re.sub(r'-\w\.(jpg|jpeg|webp)$', '-O.webp', src)
            src_lower = high_res_src.lower()
            if src_lower.endswith('.gif') or src_lower.startswith('data:'): continue
            width = img.get('width')
            height = img.get('height')
            if width and height:
                try:
                    if int(width) < 100 or int(height) < 100: continue
                except ValueError: pass
            if high_res_src not in images: images.append(high_res_src)
    if not images: raise ValueError("Nenhuma imagem encontrada na galeria do produto.")
    return images

class OpcaoCampo(NamedTuple):
    field_name: str
    options: List[str]
    col_indices: List[int]

class PlanilhaChunk:
    def __init__(self, nome, col_start, col_end):
        self.nome = nome
        self.col_start = col_start
        self.col_end = col_end
        self.campos = [] 
    def adicionar_campo(self, campo_info: dict): self.campos.append(campo_info)
    def to_dict(self): return {"nome": self.nome, "col_start": self.col_start, "col_end": self.col_end, "campos": self.campos}
    def __repr__(self): return f"<PlanilhaChunk: {self.nome} (Colunas {self.col_start}-{self.col_end}), {len(self.campos)} campos>"

EXPL_RANGE = re.compile(r"^(?:'(?P<sheet_quoted>[^']+?)'|(?P<sheet_unquoted>[^'!]+?))!(?P<start>\$?[A-Z]{1,3}\$?\d+)(?::(?P<end>\$?[A-Z]{1,3}\$?\d+))?$")
RE_DIRECT = re.compile(r"^=?\s*([A-Za-z0-9_.\-]+)\s*$")
RE_INDIRECT_SUFFIX = re.compile(r'&\s*"([A-Za-z0-9_\.]+)"')

@lru_cache(maxsize=None)
def get_s3_client():
    return boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID, aws_secret_access_key=AWS_SECRET_ACCESS_KEY, region_name=AWS_S3_REGION_NAME)

@lru_cache(maxsize=None)
def get_model(temperature=0.2):
    if not GOOGLE_API_KEY:
        raise ValueError("GOOGLE_API_KEY não está configurada. Configure a chave da API do Google Gemini no arquivo .env")
    return ChatGoogleGenerativeAI(model="gemini-1.5-flash", temperature=temperature, google_api_key=GOOGLE_API_KEY)

@lru_cache(maxsize=None)
def get_embeddings_model():
    if not GOOGLE_API_KEY:
        raise ValueError("GOOGLE_API_KEY não está configurada. Configure a chave da API do Google Gemini no arquivo .env")
    import google.generativeai as genai
    genai.configure(api_key=GOOGLE_API_KEY)
    
    return GoogleGenerativeAIEmbeddings(
        model="models/embedding-001", 
        google_api_key=GOOGLE_API_KEY,
        # Forçar modo síncrono
        transport="rest"
    )

def test_google_api_connection():
    """Testa se a conexão com a API do Google está funcionando."""
    try:
        model = get_model(temperature=0)
        response = model.invoke("Teste de conexão. Responda apenas 'OK'.")
        return True, "Conexão com Google API funcionando"
    except Exception as e:
        return False, f"Erro na conexão com Google API: {str(e)}"

class BulletPoint(BaseModel):
    bullet_point: str = Field(description="Um único bullet point, com 80 a 120 caracteres. Formato: 'TERMO EM CAIXA ALTA: Frase objetiva.'")

class AmazonListing(BaseModel):
    titulo: constr(min_length=80, max_length=120) = Field(description="Título do produto com 80 a 120 caracteres, sem símbolos como travessão ou barra.")
    bullet_points: conlist(BulletPoint, min_items=5, max_items=5) = Field(description="Uma lista com exatamente 5 bullet points.")
    descricao_produto: str = Field(description="Descrição detalhada do produto, seguindo a estrutura com subtítulos e especificações técnicas.")
    palavras_chave: str = Field(description="String única contendo 10 a 15 palavras-chave relevantes, separadas por ponto e vírgula.")

def detectar_categoria_produto(product_context: str) -> str:
    """Detecta a categoria do produto baseada no contexto fornecido."""
    context_lower = product_context.lower()
    
    # Categorias de produtos com palavras-chave específicas
    categorias = {
        'eletronicos': ['eletrônico', 'bateria', 'carregador', 'cabo', 'fone', 'smartphone', 'tablet', 'computador', 'tv', 'som'],
        'casa_jardim': ['casa', 'jardim', 'decoração', 'móvel', 'cozinha', 'banheiro', 'quarto', 'sala', 'plantas', 'ferramentas'],
        'roupas_acessorios': ['roupa', 'camisa', 'calça', 'vestido', 'sapato', 'bolsa', 'relógio', 'joia', 'óculos', 'chapéu'],
        'saude_beleza': ['saúde', 'beleza', 'cosmético', 'perfume', 'shampoo', 'creme', 'maquiagem', 'suplemento', 'vitamina'],
        'esportes': ['esporte', 'fitness', 'academia', 'corrida', 'futebol', 'tênis', 'bicicleta', 'natação', 'yoga'],
        'brinquedos': ['brinquedo', 'criança', 'boneca', 'carrinho', 'jogo', 'puzzle', 'educativo', 'infantil'],
        'automotivo': ['carro', 'auto', 'pneu', 'óleo', 'peça', 'acessório automotivo', 'motor', 'freio'],
        'livros': ['livro', 'literatura', 'romance', 'ficção', 'biografia', 'história', 'ciência', 'educação'],
        'pet': ['pet', 'cachorro', 'gato', 'ração', 'brinquedo para pet', 'coleira', 'cama para pet']
    }
    
    for categoria, palavras_chave in categorias.items():
        if any(palavra in context_lower for palavra in palavras_chave):
            return categoria
    
    return 'geral'

def gerar_palavras_chave_inteligentes(product_context: str, categoria: str) -> list:
    """Gera palavras-chave específicas baseadas na categoria do produto."""
    context_lower = product_context.lower()
    
    # Palavras-chave base por categoria
    palavras_base = {
        'eletronicos': ['eletrônicos', 'tecnologia', 'digital', 'portátil', 'wireless', 'bluetooth'],
        'casa_jardim': ['casa', 'lar', 'decoração', 'organização', 'prático', 'funcional'],
        'roupas_acessorios': ['moda', 'estilo', 'conforto', 'elegante', 'casual', 'moderno'],
        'saude_beleza': ['cuidados', 'bem-estar', 'natural', 'hidratante', 'proteção', 'tratamento'],
        'esportes': ['fitness', 'treino', 'performance', 'resistente', 'durável', 'atlético'],
        'brinquedos': ['diversão', 'educativo', 'criativo', 'seguro', 'infantil', 'desenvolvimento'],
        'automotivo': ['veículo', 'performance', 'segurança', 'manutenção', 'qualidade', 'resistente'],
        'livros': ['conhecimento', 'aprendizado', 'cultura', 'educação', 'literatura', 'informação'],
        'pet': ['animal', 'cuidado', 'conforto', 'saúde animal', 'bem-estar pet', 'qualidade'],
        'geral': ['qualidade', 'prático', 'funcional', 'durável', 'útil', 'eficiente']
    }
    
    return palavras_base.get(categoria, palavras_base['geral'])

def get_prompt_especifico_categoria(categoria: str) -> str:
    """Retorna instruções específicas baseadas na categoria do produto."""
    prompts_categoria = {
        'eletronicos': """
        INSTRUÇÕES ESPECÍFICAS PARA ELETRÔNICOS:
        - Enfatize especificações técnicas (voltagem, potência, conectividade)
        - Mencione compatibilidade com dispositivos
        - Destaque recursos de segurança e certificações
        - Inclua informações sobre garantia quando relevante
        """,
        'casa_jardim': """
        INSTRUÇÕES ESPECÍFICAS PARA CASA E JARDIM:
        - Destaque praticidade e funcionalidade no dia a dia
        - Mencione facilidade de instalação/uso
        - Enfatize durabilidade e resistência
        - Inclua informações sobre manutenção e limpeza
        """,
        'roupas_acessorios': """
        INSTRUÇÕES ESPECÍFICAS PARA ROUPAS E ACESSÓRIOS:
        - Destaque conforto e qualidade dos materiais
        - Mencione versatilidade de uso
        - Enfatize design e estilo
        - Inclua informações sobre cuidados e lavagem
        """,
        'saude_beleza': """
        INSTRUÇÕES ESPECÍFICAS PARA SAÚDE E BELEZA:
        - Enfatize benefícios para a pele/cabelo/saúde
        - Mencione ingredientes naturais quando aplicável
        - Destaque facilidade de aplicação
        - Inclua informações sobre resultados esperados
        """,
        'esportes': """
        INSTRUÇÕES ESPECÍFICAS PARA ESPORTES:
        - Destaque performance e resistência
        - Mencione conforto durante atividades
        - Enfatize durabilidade e qualidade dos materiais
        - Inclua informações sobre modalidades de uso
        """
    }
    
    return prompts_categoria.get(categoria, "")

@lru_cache(maxsize=None)
def get_main_ia_chain():
    output_parser = JsonOutputParser(pydantic_object=AmazonListing)
    prompt_template = PromptTemplate(
        input_variables=["product_context", "categoria", "instrucoes_categoria", "palavras_chave_sugeridas"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
        template="""
        Você é um especialista em criação de listings para Amazon com foco em QUALIDADE e RELEVÂNCIA.
        
        CATEGORIA DETECTADA: {categoria}
        {instrucoes_categoria}
        
        PALAVRAS-CHAVE SUGERIDAS PARA ESTA CATEGORIA: {palavras_chave_sugeridas}
        
        Crie um listing de produto para a Amazon com base nas informações fornecidas, seguindo exatamente o modelo abaixo.
        Atenção: Todo o conteúdo deve respeitar as diretrizes da Amazon Seller Central, evitando: termos proibidos como "garantido", "perfeito", "melhor", "alta qualidade", "cura" ou qualquer promessa; uso de símbolos como travessão (–), barra (/) ou emojis; declarações absolutas ou subjetivas.

        DIRETRIZES DE QUALIDADE APRIMORADAS:
        1. ANÁLISE PROFUNDA: Analise cuidadosamente todas as informações do produto antes de criar o conteúdo
        2. ESPECIFICIDADE: Use termos específicos e técnicos relevantes para a categoria
        3. BENEFÍCIOS CLAROS: Destaque benefícios práticos e tangíveis para o usuário
        4. DIFERENCIAÇÃO: Identifique e destaque características únicas do produto
        5. CONTEXTO DE USO: Descreva situações específicas onde o produto é útil

        ESTRUTURA DO LISTING

        TÍTULO (APRIMORADO)
        - Deve conter entre 80 a 120 caracteres
        - Não usar símbolos como travessão ou barra
        - Incluir nome do produto + característica principal + benefício específico
        - Usar termos de busca relevantes para a categoria
        - Exemplo: "Organizador de Mala 6 Compartimentos Impermeável para Viagem Bagagem Organizada"

        BULLET POINTS (APRIMORADOS)
        - Inserir 5 bullets, cada um com entre 80 a 120 caracteres
        - Formato: TERMO EM CAIXA ALTA: frase objetiva com benefício específico
        - Cada bullet deve abordar um aspecto diferente (funcionalidade, material, uso, manutenção, diferencial)
        - Use dados técnicos quando disponíveis (dimensões, capacidade, materiais)
        - Exemplo: ORGANIZAÇÃO INTELIGENTE: 6 compartimentos de tamanhos variados para separar roupas, sapatos e acessórios

        DESCRIÇÃO DO PRODUTO (APRIMORADA)
        - Comece com o nome do produto e seu principal benefício
        - Desenvolva 3 a 4 parágrafos estruturados:
        
        Parágrafo 1 - Funcionalidade Principal
        Explique detalhadamente como o produto funciona e seus principais recursos
        
        Parágrafo 2 - Benefícios Práticos
        Descreva os benefícios específicos para o usuário no dia a dia
        
        Parágrafo 3 - Qualidade e Durabilidade
        Destaque materiais, construção e aspectos de qualidade
        
        Especificações Técnicas (quando aplicável)
        - Dimensões exatas
        - Materiais utilizados
        - Capacidade/peso suportado
        - Certificações
        - Compatibilidade
        
        Finalize com uma frase de impacto específica para a categoria

        PALAVRAS-CHAVE (INTELIGENTES)
        - OBRIGATÓRIO: Inclua EXATAMENTE entre 12 a 15 palavras-chave relevantes
        - Use as palavras-chave sugeridas como base, mas adapte ao produto específico
        - Inclua termos de busca populares da categoria
        - Combine palavras-chave genéricas com específicas
        - Separar APENAS por ponto e vírgula (;)
        - Priorize termos que os clientes realmente pesquisam

        {format_instructions}

        **Informações Completas do Produto para Análise:**
        ```
        {product_context}
        ```
        
        IMPORTANTE: Analise profundamente as informações do produto antes de gerar o conteúdo. Seja específico, relevante e focado na categoria {categoria}.
        """
    )
    return prompt_template | get_model(temperature=0.2) | output_parser

@lru_cache(maxsize=None)
def get_extrator_chain():
    prompt_extrator = PromptTemplate.from_template("... (seu template de extrator aqui) ...")
    return prompt_extrator | get_model(temperature=0) | CommaSeparatedListOutputParser()

@lru_cache(maxsize=None)
def get_tradutor_chain():
    prompt_tradutor = PromptTemplate.from_template("... (seu template de tradutor aqui) ...")
    return prompt_tradutor | get_model(temperature=0) | StrOutputParser()

def preencher_grupo_de_colunas(ws, row: int, col_indices: list, valores: list, field_name_debug: str):
    try:
        if not valores: return
        valores_unicos = list(dict.fromkeys(valores))
        colunas_ordenadas = sorted(col_indices)
        pares_coluna_valor = list(zip(colunas_ordenadas, valores_unicos))
        for col_idx, valor_item in pares_coluna_valor:
            cell = ws.cell(row=row, column=col_idx)
            if not cell.value and valor_item and str(valor_item).lower() != 'nan':
                cell.value = valor_item
    except Exception as e:
        try: print(f"ERRO CRÍTICO DENTRO DE 'preencher_grupo_de_colunas' para o campo '{field_name_debug}': {e}")
        except Exception: print(f"ERRO CRÍTICO DENTRO DE 'preencher_grupo_de_colunas'. Falha ao imprimir detalhes do erro: {e}")

ia_cache = {}
def obter_com_cache(key, func, *args, **kwargs):
    if key in ia_cache: return ia_cache[key]
    resultado = func(*args, **kwargs)
    ia_cache[key] = resultado
    return resultado

def limpar_cache_ia():
    """Limpa o cache manual da IA para evitar problemas com preenchimento de campos."""
    global ia_cache
    ia_cache.clear()
    print("Cache da IA limpo com sucesso.")

def validar_qualidade_conteudo(conteudo_gerado: dict, categoria: str) -> dict:
    """Valida a qualidade do conteúdo gerado e retorna um score de qualidade."""
    score = 0
    feedback = []
    max_score = 100
    
    # Validação do título (25 pontos)
    titulo = conteudo_gerado.get('titulo', '')
    if titulo:
        if 80 <= len(titulo) <= 120:
            score += 15
        else:
            feedback.append(f"Título com {len(titulo)} caracteres (ideal: 80-120)")
        
        # Verifica se contém termos específicos da categoria
        if categoria != 'geral':
            categorias_termos = {
                'eletronicos': ['eletrônico', 'digital', 'tecnologia'],
                'casa_jardim': ['casa', 'jardim', 'organização'],
                'roupas_acessorios': ['moda', 'estilo', 'conforto'],
                'saude_beleza': ['beleza', 'cuidado', 'bem-estar'],
                'esportes': ['fitness', 'esporte', 'treino']
            }
            termos_categoria = categorias_termos.get(categoria, [])
            if any(termo in titulo.lower() for termo in termos_categoria):
                score += 10
            else:
                feedback.append("Título não contém termos específicos da categoria")
    else:
        feedback.append("Título ausente")
    
    # Validação dos bullet points (25 pontos)
    bullet_points = conteudo_gerado.get('bullet_points', [])
    if bullet_points and len(bullet_points) == 5:
        score += 10
        bullets_validos = 0
        for bp in bullet_points:
            bullet_text = bp.get('bullet_point', '') if isinstance(bp, dict) else str(bp)
            if 80 <= len(bullet_text) <= 120 and ':' in bullet_text:
                bullets_validos += 1
        score += (bullets_validos / 5) * 15
        if bullets_validos < 5:
            feedback.append(f"Apenas {bullets_validos}/5 bullet points com formato correto")
    else:
        feedback.append(f"Número incorreto de bullet points: {len(bullet_points) if bullet_points else 0}")
    
    # Validação da descrição (25 pontos)
    descricao = conteudo_gerado.get('descricao_produto', '')
    if descricao:
        if len(descricao) >= 200:
            score += 15
        else:
            feedback.append(f"Descrição muito curta: {len(descricao)} caracteres")
        
        # Verifica estrutura (parágrafos)
        paragrafos = descricao.split('\n\n')
        if len(paragrafos) >= 3:
            score += 10
        else:
            feedback.append("Descrição sem estrutura adequada de parágrafos")
    else:
        feedback.append("Descrição ausente")
    
    # Validação das palavras-chave (25 pontos)
    palavras_chave = conteudo_gerado.get('palavras_chave', '')
    if palavras_chave:
        keywords_list = [kw.strip() for kw in palavras_chave.split(';') if kw.strip()]
        if 12 <= len(keywords_list) <= 15:
            score += 15
        else:
            feedback.append(f"Número de palavras-chave inadequado: {len(keywords_list)}")
        
        # Verifica se não há vírgulas (erro comum)
        if ',' not in palavras_chave:
            score += 10
        else:
            feedback.append("Palavras-chave contêm vírgulas (deve usar apenas ponto e vírgula)")
    else:
        feedback.append("Palavras-chave ausentes")
    
    return {
        'score': min(score, max_score),
        'qualidade': 'Excelente' if score >= 90 else 'Boa' if score >= 70 else 'Regular' if score >= 50 else 'Ruim',
        'feedback': feedback
    }

def priorizar_campos_inteligente(product_data: dict, categoria: str) -> list:
    """Prioriza campos para preenchimento baseado na categoria e dados disponíveis."""
    # Campos essenciais sempre priorizados
    campos_essenciais = [
        'titulo', 'descricao_produto', 'bullet_points', 'palavras_chave',
        'marca', 'modelo', 'cor_principal'
    ]
    
    # Campos específicos por categoria
    campos_categoria = {
        'eletronicos': [
            'voltagem', 'potencia', 'conectividade', 'compatibilidade',
            'certificacao', 'garantia', 'tipo_bateria'
        ],
        'casa_jardim': [
            'material_principal', 'dimensoes', 'peso_maximo_suportado',
            'facilidade_instalacao', 'resistencia_agua', 'manutencao'
        ],
        'roupas_acessorios': [
            'tamanho', 'material_tecido', 'tipo_fechamento', 'ocasiao_uso',
            'cuidados_lavagem', 'estilo', 'genero'
        ],
        'saude_beleza': [
            'ingredientes_principais', 'tipo_pele', 'beneficios',
            'modo_uso', 'dermatologicamente_testado', 'natural_organico'
        ],
        'esportes': [
            'modalidade_esportiva', 'nivel_usuario', 'material_resistente',
            'tecnologia_performance', 'tamanho_peso', 'durabilidade'
        ]
    }
    
    # Combina campos essenciais com específicos da categoria
    campos_priorizados = campos_essenciais.copy()
    if categoria in campos_categoria:
        campos_priorizados.extend(campos_categoria[categoria])
    
    # Adiciona campos que têm dados disponíveis no produto
    campos_com_dados = []
    for campo in product_data.keys():
        if product_data[campo] and str(product_data[campo]).strip() and str(product_data[campo]).lower() != 'nan':
            if campo not in campos_priorizados:
                campos_com_dados.append(campo)
    
    # Retorna lista priorizada
    return campos_priorizados + campos_com_dados[:10]  # Limita campos extras

def detectar_codificacao(path: str) -> str:
    with open(path, "rb") as f: bruto = f.read(10_000)
    return chardet.detect(bruto)["encoding"] or "utf-8"

def carregar_docs_csv(path: str, delimiter: str = ","):
    if not os.path.exists(path): return []
    loader = CSVLoader(file_path=path, csv_args={"delimiter": delimiter, "quotechar": '"'}, encoding=detectar_codificacao(path))
    return loader.load()

def carregar_explorar_dados_csv(path: str) -> List[Document]:
    if not os.path.exists(path): return []
    try: encoding = detectar_codificacao(path); df = pd.read_csv(path, encoding=encoding)
    except Exception as e: return []
    documents = []
    for index, row in df.iterrows():
        node_number = row.get("Número do Node"); browse_path = row.get("Caminho de Navegação"); positive_keywords = row.get("Palavras-Chave Positivas"); negative_keywords = row.get("Palavras-Chave Negativas")
        if pd.notna(node_number) and pd.notna(browse_path):
            content_parts = [f"Caminho de Navegação: {browse_path}."]
            if pd.notna(positive_keywords): content_parts.append(f"Tópicos e funções relevantes para este caminho: {positive_keywords}.")
            if pd.notna(negative_keywords): content_parts.append(f"Tópicos e funções a evitar para este caminho: {negative_keywords}.")
            enriched_page_content = " ".join(content_parts); full_path_with_id = f"{browse_path} ({node_number})"
            doc = Document(page_content=enriched_page_content, metadata={"tipo_dado": "caminho_navegacao", "Número do Node": str(node_number).strip(), "Caminho de Navegação": str(browse_path).strip(), "valor_opcao": full_path_with_id, "positive_keywords": str(positive_keywords).strip() if pd.notna(positive_keywords) else "", "negative_keywords": str(negative_keywords).strip() if pd.notna(negative_keywords) else ""})
            documents.append(doc)
    return documents

def carregar_tipos_de_produto(path: str) -> List[Document]:
    if not os.path.exists(path): return []
    try: encoding = detectar_codificacao(path); df = pd.read_csv(path, encoding=encoding)
    except Exception as e: return []
    df_tipos_produto = df[df['campo'] == 'Tipo de produto -'].copy(); documentos = []
    for index, row in df_tipos_produto.iterrows():
        tipo_produto_valor = row.get("valor_opcao")
        if pd.notna(tipo_produto_valor):
            termo_em_pt = tipo_produto_valor.replace('_', ' ').lower()
            conteudo_descritivo = f"Tipo de Produto da Amazon: {tipo_produto_valor}. Categoria de produto para o item {termo_em_pt}."
            doc = Document(page_content=conteudo_descritivo, metadata={"tipo_dado": "tipo_de_produto", "valor_opcao": tipo_produto_valor})
            documentos.append(doc)
    return documentos

@lru_cache(maxsize=None)
def get_vectorstore():
    embeddings = get_embeddings_model()
    faiss_index_path = os.path.join(settings.BASE_DIR, "faiss_index")
    if os.path.exists(os.path.join(faiss_index_path, "index.faiss")):
        return FAISS.load_local(faiss_index_path, embeddings, allow_dangerous_deserialization=True)
    j_csv_path = os.path.join(settings.BASE_DIR, "memo", NOME_CSV_J); j_planilha_csv_path = os.path.join(settings.BASE_DIR, "memo", NOME_CSV_J_PLANILHA); valores_validos_csv_path = os.path.join(settings.BASE_DIR, "memo", NOME_CSV_VALORES_VALIDOS) ; explorar_dados_csv_path = os.path.join(settings.BASE_DIR, "memo", NOME_CSV_EXPLORAR_DADOS); definicoes_dados_csv_path = os.path.join(settings.BASE_DIR, "memo", NOME_CSV_DEFINICOES_DADOS)
    all_docs = (carregar_docs_csv(j_csv_path) + carregar_docs_csv(j_planilha_csv_path) + carregar_docs_csv(valores_validos_csv_path) + carregar_explorar_dados_csv(explorar_dados_csv_path) + carregar_docs_csv(definicoes_dados_csv_path))
    if not all_docs: return None
    vectorstore = FAISS.from_documents(all_docs, embeddings)
    if not os.path.exists(faiss_index_path): os.makedirs(faiss_index_path)
    vectorstore.save_local(faiss_index_path, index_name="index")
    return vectorstore

def escolher_com_ia(product_data: dict, fields_to_fill: List[Dict], assumed_items: FrozenSet[Tuple[str, str]], retriever) -> Tuple[Dict[str, str], int]:
    if not fields_to_fill:
        return {}, 0
    
    titulo_produto = product_data.get("titulo", "produto")
    product_context_str = format_product_context(product_data)
    
    assumed_values = dict(assumed_items)
    context_query = f"Forneça informações e especificações para o produto '{titulo_produto}' para ajudar no preenchimento de seus atributos."
    
    # Implementar retry para resolver problemas de conexão
    max_retries = 3
    retry_delay = 2
    relevant_docs = []
    
    for attempt in range(max_retries):
        try:
            relevant_docs = retriever.invoke(context_query)
            break
        except Exception as e:
            print(f"Tentativa {attempt + 1} falhou ao buscar documentos relevantes: {e}")
            if attempt < max_retries - 1:
                print(f"Tentando novamente em {retry_delay} segundos...")
                import time
                time.sleep(retry_delay)
                retry_delay *= 2  # Backoff exponencial
            else:
                print("Todas as tentativas falharam. Continuando sem documentos relevantes.")
                relevant_docs = []
    
    retriever_context_str = "\n".join([f"- {doc.page_content}" for doc in relevant_docs]) or "Nenhuma informação adicional encontrada na base de conhecimento."
    
    full_context = f"DADOS DO PRODUTO:\n{product_context_str}\n\nINFORMAÇÕES ADICIONAIS DA BASE DE CONHECIMENTO:\n{retriever_context_str}"
    
    fields_json_str = json.dumps(fields_to_fill, indent=2, ensure_ascii=False)
    
    multi_value_field_names = [f['field_name'] for f in fields_to_fill if f.get('multi_value', False)]
    multi_value_context = (f"Para os seguintes campos, se aplicável, retorne uma LISTA JSON de strings com os valores relevantes: {', '.join(multi_value_field_names)}\n" if multi_value_field_names else "")
    
    critical_field_names = [f['field_name'] for f in fields_to_fill if f.get('is_critical', False)]
    critical_context = ""
    if critical_field_names:
        critical_context = (
            "ATENÇÃO MÁXIMA: Os seguintes campos são CRÍTICOS e devem ter um valor válido e preciso: "
            f"**{', '.join(critical_field_names)}**. "
            "É OBRIGATÓRIO que você escolha a melhor opção da lista para estes campos. NÃO use 'nan' para eles, a menos que seja absolutamente impossível determinar um valor.\n"
        )
    
    prompt = PromptTemplate(
        input_variables=["product", "context", "assumed_values", "fields_json", "multi_value_context", "critical_context"],
        template=(
            "Você é um especialista em catalogação de produtos para e-commerce, seguindo as diretrizes da Amazon.\n"
            "Sua tarefa é preencher vários campos para um produto com base nas opções disponíveis para cada um.\n"
            "Analise o título do produto, o CONTEXTO COMPLETO e os valores de referência para tomar a decisão mais precisa para CADA campo.\n\n"
            "REGRAS IMPORTANTES:\n"
            "- {critical_context}"
            "- Para a maioria dos campos, retorne uma única string como valor.\n"
            "- {multi_value_context}"
            "- CAMPOS DE UNIDADE: NUNCA preencha automaticamente os seguintes campos de unidade específicos: 'Unidade de altura', 'Unidade de comprimento', 'Unidade da largura', 'Unidade de Altura do Pacote Principal', 'Unidade de Comprimento do Pacote Principal', 'Unidade de Largura do Pacote Principal', 'Unidade de Peso do Pacote Principal', 'Unidade da profundidade do artigo', 'Unidade de altura do artigo', 'Unidade de largura do artigo'. Para estes campos, sempre use 'nan'. Para outros campos de unidade (como 'Unidade de comprimento do item', 'Unidade de largura do item', 'Unidade de altura do item'), preencha apenas se explicitamente fornecido no contexto.\n"
            "- Se nenhuma opção for adequada ou se a informação for desconhecida, use a string 'nan'.\n\n"
            "--- DADOS DO PRODUTO E CONTEXTO ---\nProduto de Referência: '{product}'\n\nContexto Completo:\n{context}\n\n"
            "--- VALORES DE REFERÊNCIA ---\n{assumed_values}\n\n"
            "--- CAMPOS PARA PREENCHER ---\n{fields_json}\n\n"
            "--- INSTRUÇÃO DE SAÍDA ---\n"
            "Responda APENAS com um único objeto JSON válido, mapeando cada 'field_name' para a sua escolha.\n"
            "Exemplo de formato de saída: {{\"Material\": [\"Plástico\", \"Metal\"], \"Cor\": \"Azul\", \"Regulamentações...\": \"Não aplicável\"}}\n"
            "Não inclua NENHUM texto, explicação ou formatação de código antes ou depois do objeto JSON."
        )
    ).format(
        product=titulo_produto,
        context=full_context,
        assumed_values=json.dumps(assumed_values, indent=2, ensure_ascii=False),
        fields_json=fields_json_str,
        multi_value_context=multi_value_context,
        critical_context=critical_context
    )
    # Usar cache inteligente para a resposta da IA
    def call_ai_function():
        response = get_model(temperature=0.2).invoke(prompt)
        response_content = response.content.strip()
        total_tokens = response.response_metadata.get('usage', {}).get('total_tokens', 0)
        
        try:
            json_match = re.search(r'\{.*\}', response_content, re.DOTALL)
            if json_match:
                ai_choices = json.loads(json_match.group(0))
                return {'choices': ai_choices, 'tokens': total_tokens}
            return {'choices': {}, 'tokens': total_tokens}
        except json.JSONDecodeError:
            return {'choices': {}, 'tokens': total_tokens}
    
    # Contexto para cache (sem incluir o prompt completo para economizar espaço)
    cache_context = {
        'product_title': titulo_produto,
        'fields_count': len(fields_for_ai_batch),
        'field_names': [f['field_name'] for f in fields_for_ai_batch],
        'assumed_values': assumed_values
    }
    
    cached_result = get_or_cache_ai_response(
        prompt=prompt[:500] + "...",  # Usar apenas início do prompt para cache
        context=cache_context,
        ai_function=call_ai_function
    )
    
    return cached_result['choices'], cached_result['tokens']

def processar_chunk_com_ia(chunk_name: str, chunk_data: dict, product_data: dict, retriever_context_info: str, campos_criticos: set, persona_especialista: str, force_regenerate: bool = False):
    print(f"\nINFO: Processando o Chunk '{chunk_name}'...")
    
    titulo_produto = product_data.get("titulo", "produto")
    
    # Verificar se o produto já tem dados na memória para este chunk
    if not force_regenerate:
        try:
            identifier = extract_product_identifier(product_data)
            cached_data = check_product_in_memory(product_data)
            
            if cached_data[0] and cached_data[1]:
                saved_content = cached_data[1].get('generated_content', {})
                
                # Verificar se há dados relevantes para este chunk
                chunk_fields = [campo['cabecalho_l5'] for campo in chunk_data.get('campos', [])]
                cached_chunk_data = {}
                
                for field in chunk_fields:
                    if field in saved_content.get('outros_campos', {}):
                        cached_chunk_data[field] = saved_content['outros_campos'][field]
                
                if cached_chunk_data:
                    print(f"INFO: Reutilizando dados da memória para chunk '{chunk_name}' do produto {identifier}")
                    return cached_chunk_data
                else:
                    print(f"INFO: Produto {identifier} encontrado na memória, mas sem dados para chunk '{chunk_name}'")
        except Exception as e:
            print(f"AVISO: Erro ao verificar memória para chunk '{chunk_name}': {e}")
    
    product_context_str = format_product_context(product_data)
    full_context = f"DADOS COMPLETOS DO PRODUTO:\n{product_context_str}\n\nINFORMAÇÕES ADICIONAIS DA BASE DE CONHECIMENTO:\n{retriever_context_info}"

    chunk = PlanilhaChunk(chunk_data['nome'], chunk_data['col_start'], chunk_data['col_end'])
    chunk.campos = chunk_data['campos']

    campos_vazios = chunk.campos
    if not campos_vazios:
        print(f"INFO: Nenhum campo a ser preenchido no chunk '{chunk_name}'. Pulando.")
        return {}

    nomes_l4_dos_campos_vazios = list(dict.fromkeys([campo['cabecalho_l4'] for campo in campos_vazios if campo['cabecalho_l4']]))
    if not nomes_l4_dos_campos_vazios:
        print(f"INFO: Nenhum campo com cabeçalho na linha 4 para ser avaliado no chunk '{chunk_name}'.")
        return {}

    print(f"  -> Estágio 1: Verificando a relevância de {len(nomes_l4_dos_campos_vazios)} grupos de campos para o produto...")
    prompt_triagem = (
        f"Com base nas informações completas do produto: \n\n{product_context_str}\n\n"
        f"Avalie a lista de grupos de atributos a seguir: {', '.join(nomes_l4_dos_campos_vazios)}. "
        f"Responda APENAS com um objeto JSON contendo uma única chave 'campos_relevantes', "
        f"cujo valor é uma lista de strings com os nomes dos atributos da lista que são REALMENTE RELEVANTES. "
        f"Exemplo: para uma VELA, o atributo 'Voltagem' é irrelevante e não deve ser incluído na resposta."
    )
    nomes_l4_relevantes = []
    try:
        # Usar cache inteligente para triagem
        def call_triagem_ai():
            modelo_triagem = get_model(temperature=0.2)
            resposta_triagem_raw = modelo_triagem.invoke(prompt_triagem).content.strip()
            json_match_triagem = re.search(r'\{.*\}', resposta_triagem_raw, re.DOTALL)
            return json.loads(json_match_triagem.group(0))
        
        triagem_context = {
            'product_title': titulo_produto,
            'campos_count': len(nomes_l4_dos_campos_vazios),
            'campos_names': nomes_l4_dos_campos_vazios
        }
        
        triagem_result = get_or_cache_ai_response(
            prompt=prompt_triagem,
            context=triagem_context,
            ai_function=call_triagem_ai
        )
        
        nomes_l4_relevantes = triagem_result['campos_relevantes']
        print(f"  -> Triagem concluída. Grupos de campos relevantes identificados: {nomes_l4_relevantes}")
    except (AttributeError, json.JSONDecodeError, KeyError) as e:
        print(f"  -> AVISO: Não foi possível determinar os campos relevantes na triagem ({e}). O preenchimento seguirá sem o filtro de relevância.")
        nomes_l4_relevantes = nomes_l4_dos_campos_vazios

    print(f"  -> Aplicando 'Passe VIP' para campos críticos...")
    for campo_critico in campos_criticos:
        if campo_critico not in nomes_l4_relevantes:
            if any(campo['cabecalho_l4'] == campo_critico for campo in campos_vazios):
                print(f"    -> PASSE VIP: Campo crítico '{campo_critico}' não foi escolhido pela IA. Adicionando manualmente à lista.")
                nomes_l4_relevantes.append(campo_critico)

    campos_para_preencher = [campo for campo in campos_vazios if campo['cabecalho_l4'] in nomes_l4_relevantes]
    if not campos_para_preencher:
        print(f"INFO: Nenhum campo relevante encontrado para preencher no chunk '{chunk_name}' após a triagem.")
        return {}

    print(f"  -> Estágio 2: Pedindo à IA especialista para preencher os {len(campos_para_preencher)} campos relevantes...")
    nomes_campos_criticos = [campo['cabecalho_l5'] for campo in campos_para_preencher if campo['cabecalho_l4'] in campos_criticos]
    contexto_critico = ""
    if nomes_campos_criticos:
        contexto_critico = (f"ATENÇÃO MÁXIMA: Os seguintes campos deste grupo são CRÍTICOS e devem ter um valor válido: "
                            f"**{', '.join(nomes_campos_criticos)}**. Evite 'nan' para eles a todo custo.\n")

    campos_str_detalhado = "\n".join([f"- **{campo['cabecalho_l5']}** (do grupo '{campo['cabecalho_l4']}')" for campo in campos_para_preencher])

    prompt_final = (
        "{persona}\n\n"
        "Sua missão é preencher DE FORMA PRECISA E COMPLETA **CADA CAMPO** da lista fornecida...\n\n"
        "### REGRAS DE OURO:\n"
        "1.  **NÃO INVENTE INFORMAÇÕES:** Use apenas dados fornecidos no contexto do produto. Se não houver informação suficiente, use 'nan'.\n"
        "2.  **PENSE PASSO A PASSO:** Analise cada campo individualmente com base no contexto fornecido.\n"
        "3.  **SAÍDA EXCLUSIVAMENTE EM JSON:** Responda apenas com um objeto JSON válido, sem texto adicional.\n"
        "4.  **CHAVES EXATAS:** Use exatamente os nomes dos campos fornecidos como chaves no JSON.\n"
        "5.  **VALORES EM PORTUGUÊS BRASILEIRO.**\n"
        "6.  **CAMPOS DE UNIDADE:** NUNCA preencha automaticamente os seguintes campos de unidade específicos: 'Unidade de altura', 'Unidade de comprimento', 'Unidade da largura', 'Unidade de Altura do Pacote Principal', 'Unidade de Comprimento do Pacote Principal', 'Unidade de Largura do Pacote Principal', 'Unidade de Peso do Pacote Principal', 'Unidade da profundidade do artigo', 'Unidade de altura do artigo', 'Unidade de largura do artigo'. Para estes campos, sempre use 'nan'. Para outros campos de unidade (como 'Unidade de comprimento do item', 'Unidade de largura do item', 'Unidade de altura do item'), preencha apenas se explicitamente fornecido no contexto.\n"
        "7.  **CAMPOS CRÍTICOS:** {contexto_critico}\n\n"
        "### DADOS PARA ANÁLISE\n"
        "**Produto de Referência:**\n`{titulo_produto}`\n\n"
        "**Contexto Completo:**\n{context_info}\n\n"
        "### CAMPOS PARA PREENCHER (APENAS DESTE GRUPO RELEVANTE)\n"
        "{campos_str_detalhado}\n"
        "### INSTRUÇÃO DE SAÍDA FINAL\n"
        "Gere o objeto JSON com uma entrada para CADA um dos campos listados acima. Não omita nenhum."
    ).format(
        persona=persona_especialista,
        contexto_critico=contexto_critico or "Nenhum campo crítico neste grupo.",
        titulo_produto=titulo_produto,
        context_info=full_context,
        campos_str_detalhado=campos_str_detalhado
    )
    
    try:
        # Usar cache inteligente para preenchimento final
        def call_final_ai():
            resposta_ia = get_model(temperature=0.2).invoke(prompt_final)
            response_content = resposta_ia.content.strip()
            json_match = re.search(r'\{.*\}', response_content, re.DOTALL)
            if not json_match:
                return {}
            return json.loads(json_match.group(0))
        
        final_context = {
            'product_title': titulo_produto,
            'chunk_name': chunk_name,
            'campos_count': len(campos_para_preencher),
            'campos_names': [campo['cabecalho_l5'] for campo in campos_para_preencher],
            'campos_criticos': list(campos_criticos)
        }
        
        ai_choices = get_or_cache_ai_response(
            prompt=prompt_final[:1000] + "...",  # Usar apenas início do prompt para cache
            context=final_context,
            ai_function=call_final_ai
        )
        
        if not ai_choices:
            print(f"AVISO: Nenhum resultado válido da IA para o chunk '{chunk_name}'.")
            return {}
        
        # Salvar resultados do chunk na memória
        try:
            identifier = extract_product_identifier(product_data)
            existing_data = check_product_in_memory(product_data)
            
            if existing_data[0] and existing_data[1]:
                # Atualizar dados existentes
                saved_content = existing_data[1].get('generated_content', {})
                if 'outros_campos' not in saved_content:
                    saved_content['outros_campos'] = {}
                
                # Adicionar novos campos do chunk
                saved_content['outros_campos'].update(ai_choices)
                
                # Salvar dados atualizados
                save_generated_content_to_memory(
                    product_data=product_data,
                    generated_content=saved_content,
                    force_update=True
                )
                print(f"INFO: Dados do chunk '{chunk_name}' salvos na memória para produto {identifier}")
            else:
                # Criar nova entrada na memória
                new_content = {'outros_campos': ai_choices}
                save_generated_content_to_memory(
                    product_data=product_data,
                    generated_content=new_content,
                    force_update=False
                )
                print(f"INFO: Nova entrada na memória criada para produto {identifier} com dados do chunk '{chunk_name}'")
        except Exception as e:
            print(f"AVISO: Erro ao salvar chunk '{chunk_name}' na memória: {e}")
        
        return ai_choices

    except json.JSONDecodeError:
        print(f"ERRO: Falha ao decodificar o JSON para o chunk '{chunk_name}'. Resposta: {response_content}")
        return {}
    except Exception as e:
        print(f"ERRO inesperado ao processar o chunk '{chunk_name}': {e}")
        return {}

def filtrar_documentos_dinamicamente(documentos_brutos: list, titulo_produto: str) -> list:
    documentos_filtrados = []
    for doc in documentos_brutos:
        palavras_negativas_str = doc.metadata.get('negative_keywords', '')
        if not palavras_negativas_str: documentos_filtrados.append(doc); continue
        lista_palavras_negativas = [kw.strip() for kw in palavras_negativas_str.split(',')]
        houve_conflito = False
        for palavra in lista_palavras_negativas:
            if palavra and palavra in titulo_produto: houve_conflito = True; break
        if not houve_conflito: documentos_filtrados.append(doc)
    return documentos_filtrados

def filtrar_por_relevancia_lexical(documentos: list, must_have_keywords: list) -> list:
    if not must_have_keywords: return documentos
    documentos_filtrados = []
    for doc in documentos:
        conteudo_doc_lower = doc.page_content.lower()
        if any(keyword.strip() in conteudo_doc_lower for keyword in must_have_keywords):
            documentos_filtrados.append(doc)
    return documentos_filtrados

def parse_form_data(post_data, files_data):
    output = {}; key_regex = re.compile(r'\[([^\]]*)\]')
    for key, value in post_data.items():
        parts = key_regex.findall(key)
        if not parts: continue
        main_key = key.split('[')[0]
        current_level = output.setdefault(main_key, {}).setdefault(parts[0], {})
        inner_key = parts[1]
        current_level[inner_key] = value
    for key, uploaded_file in files_data.items():
        parts = key_regex.findall(key)
        if not parts: continue
        main_key, index, field, sub_field = key.split('[')[0], parts[0], parts[1], parts[2].replace(']', '')
        output.setdefault(main_key, {}).setdefault(index, {}).setdefault(field, {})[sub_field] = uploaded_file.name
    return output

def mapear_chunks_da_planilha(ws) -> dict[str, PlanilhaChunk]:
    chunks = {}
    for merged_cell_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        if min_row == 3:
            chunk_name = ws.cell(row=min_row, column=min_col).value
            if not chunk_name: continue
            chunk_obj = PlanilhaChunk(nome=chunk_name, col_start=min_col, col_end=max_col)
            for col in range(min_col, max_col + 1):
                header_l4 = str(ws.cell(row=4, column=col).value or '').strip()
                header_l5 = str(ws.cell(row=5, column=col).value or '').strip()
                if header_l4 or header_l5: chunk_obj.adicionar_campo({'col': col, 'cabecalho_l4': header_l4, 'cabecalho_l5': header_l5})
            chunks[chunk_name] = chunk_obj
    return chunks

def _ler_range(ws, coord):
    bounds = coord if ':' in coord else f"{coord}:{coord}"; min_col, min_row, max_col, max_row = range_boundaries(bounds); valores = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            valor = ws.cell(row=r, column=c).value
            if valor is not None: valores.append(str(valor))
    return valores

def construir_mapas_nomeados(wb):
    full_map = {}; suffix_map = {}
    for name, defn in wb.defined_names.items():
        if not defn.destinations: continue
        for sheet_title, coord in defn.destinations:
            try:
                ws = wb[sheet_title]; full_map[name] = (ws, coord); suffix = name.split('_', 1)[-1]; suffix_map.setdefault(suffix, []).append(name)
            except KeyError: print(f"AVISO: Planilha '{sheet_title}' não encontrada para o range nomeado '{name}'.")
    return full_map, suffix_map

def extrair_opcoes(wb, ws, bruto_formula, nr_full_map, nr_suffix_map, row):
    f_bruto = bruto_formula or ''
    if f_bruto.strip().upper().startswith('IF(') and 'INDIRECT' in f_bruto.upper():
        inner_indirects = re.findall(r'INDIRECT\((.*?)\)', f_bruto, flags=re.IGNORECASE)
        for inner in inner_indirects:
            # *** INÍCIO DA CORREÇÃO: Garante que 'opts' sempre tenha um valor ***
            opts = extrair_opcoes(wb, ws, f"INDIRECT({inner})", nr_full_map, nr_suffix_map, row)
            if opts: return opts
            # *** FIM DA CORREÇÃO ***
    
    # *** INÍCIO DA CORREÇÃO: Inicializa 'opts' antes do loop ***
    opts = []
    literal_names = re.findall(r'"([A-Za-z0-9_\.\[\]\=\#\-]+)"', f_bruto)
    for name in literal_names:
        if name in nr_full_map: 
            ws_nr, coord = nr_full_map[name]
            opts = _ler_range(ws_nr, coord)
            if opts: return opts
    # *** FIM DA CORREÇÃO ***

    f = f_bruto.lstrip('=').strip(); m = RE_DIRECT.match(f)
    if m:
        name = m.group(1)
        if name in nr_full_map: 
            ws_nr, coord = nr_full_map[name]
            opts = _ler_range(ws_nr, coord)
            if opts: return opts
    if f.upper().startswith('INDIRECT'):
        prefix = str(ws.cell(row=row, column=3).value or '').replace('-', '_').replace(' ', '_'); cleaned = re.sub(r'VLOOKUP\([^)]*\)', '', f_bruto); match = RE_INDIRECT_SUFFIX.search(cleaned)
        if match:
            suffix = match.group(1); cand = prefix + suffix
            if cand in nr_full_map: 
                ws1, coord1 = nr_full_map[cand]
                opts = _ler_range(ws1, coord1)
                if opts: return opts
            for nm in nr_full_map:
                if suffix in nm: 
                    ws2, coord2 = nr_full_map[nm]
                    opts = _ler_range(ws2, coord2)
                    if opts: return opts
    if f.startswith('"') and f.endswith('"'): return [valor.strip() for valor in re.split(r'[;,]', f.strip('"')) if valor.strip()]
    if f.startswith('{') and f.endswith('}'): return [valor.strip() for valor in re.split(r'[;,]', f.strip('{}')) if valor.strip()]
    m2 = EXPL_RANGE.match(f)
    if m2:
        sheet = (m2.group('sheet_quoted') or m2.group('sheet_unquoted') or ws.title).strip(); coord = m2.group('start') + (':' + m2.group('end') if m2.group('end') else '')
        try: 
            tgt = wb[sheet]
            opts = _ler_range(tgt, coord)
            if opts: return opts
        except KeyError: print(f"AVISO: Planilha referenciada '{sheet}' não encontrada: {f_bruto}")
    if f in nr_full_map: 
        ws_fb, coord_fb = nr_full_map[f]
        opts = _ler_range(ws_fb, coord_fb)
        if opts: return opts
    return []

def coletar_opcoes_campo(wb, sheet_name: str, data_row: int) -> dict:
    ws = wb[sheet_name]; nr_full_map, nr_suffix_map = construir_mapas_nomeados(wb); field_groups = {}
    for dv in ws.data_validations.dataValidation:
        if dv and (dv.type or '').lower() != 'list': continue
        opts = extrair_opcoes(wb, ws, dv.formula1 or "", nr_full_map, nr_suffix_map, data_row)
        if not opts: opts = ["nan"]
        cols = []
        for cr in dv.ranges:
            min_c, min_r, max_c, max_r = cr.bounds
            if min_r <= data_row <= max_r: cols.extend(range(min_c, max_c + 1))
        if not cols: continue
        cabecalho = str(ws.cell(row=4, column=cols[0]).value or "").strip()
        if not cabecalho: continue
        if cabecalho not in field_groups: field_groups[cabecalho] = {'field_name': cabecalho, 'options': opts, 'col_indices': cols}
        else: field_groups[cabecalho]['col_indices'].extend(cols)
    return field_groups

def extract_template_fields(template_path: str):
    if not os.path.exists(template_path): raise FileNotFoundError(f"O arquivo de template não foi encontrado em: {template_path}")
    workbook = load_workbook(filename=template_path, read_only=True, keep_vba=True)
    try: sheet = workbook["Modelo"]
    except KeyError: raise ValueError("A aba 'Modelo' não foi encontrada no arquivo.")
    headers_map = {}; max_col = sheet.max_column
    for col_idx in range(1, max_col + 1):
        header_l4 = str(sheet.cell(row=4, column=col_idx).value or '').strip()
        header_l5 = str(sheet.cell(row=5, column=col_idx).value or '').strip()
        if not header_l4: continue
        if header_l4 not in headers_map: headers_map[header_l4] = {'tech_names': []}
        headers_map[header_l4]['tech_names'].append({'name': header_l5, 'col': col_idx})
    template_structure = {}
    for header, data in headers_map.items():
        if len(data['tech_names']) > 1: template_structure[header] = { item['name']: "" for item in data['tech_names'] if item['name'] }
        else:
            if data['tech_names']: template_structure[header] = ""
            else: template_structure[header] = ""
    return template_structure, headers_map

def parse_ai_response_to_dict(text_response: str) -> dict:
    response_dict = {}; lines = text_response.strip().split('\n')
    for line in lines:
        if '::' not in line: continue
        key, value = line.split('::', 1); key = key.strip(); value = value.strip()
        if '>' in key:
            parent_key, child_key = key.split('>', 1); parent_key = parent_key.strip(); child_key = child_key.strip()
            if parent_key not in response_dict: response_dict[parent_key] = {}
            response_dict[parent_key][child_key] = value
        else: response_dict[key] = value
    return response_dict

def write_to_sheet(sheet, row_num, ai_response_dict, headers_map):
    for header, value in ai_response_dict.items():
        if header not in headers_map: continue
        if isinstance(value, dict):
            for tech_name, sub_value in value.items():
                for item in headers_map[header]['tech_names']:
                    if item['name'] == tech_name: sheet.cell(row=row_num, column=item['col'], value=sub_value); break
        else:
            if headers_map[header]['tech_names']: col_idx = headers_map[header]['tech_names'][0]['col']; sheet.cell(row=row_num, column=col_idx, value=value)

def identificar_campos_multi_valor(ws) -> set:
    base_name_counts = defaultdict(int); base_name_to_header_l4 = {}; regex_base_name = re.compile(r"^(.*?)(?:\[.*?\])?#\d+\.value$"); header_l4_cache = {}
    for col_idx in range(1, ws.max_column + 1):
        header_l5 = ws.cell(row=5, column=col_idx).value
        if not header_l5 or not isinstance(header_l5, str): continue
        match = regex_base_name.match(header_l5)
        if match:
            base_name = match.group(1); base_name_counts[base_name] += 1
            if base_name not in base_name_to_header_l4:
                current_header_l4 = ""
                for back_col_idx in range(col_idx, 0, -1):
                    cached_val = header_l4_cache.get(back_col_idx)
                    if cached_val: current_header_l4 = cached_val; break
                    cell_val = ws.cell(row=4, column=back_col_idx).value
                    if cell_val: current_header_l4 = str(cell_val).strip(); header_l4_cache[back_col_idx] = current_header_l4; break
                if current_header_l4: base_name_to_header_l4[base_name] = current_header_l4
    multi_value_fields = set()
    for base_name, count in base_name_counts.items():
        if count > 1:
            if base_name in base_name_to_header_l4: header_l4 = base_name_to_header_l4[base_name]; multi_value_fields.add(header_l4)
    return multi_value_fields

def normalizar_ncm(bruto_ncm: str) -> str: return re.sub(r'\D', '', bruto_ncm or '') or ""
def gerar_ncm_aleatorio() -> str:
    while True:
        ncm = f"{random.randint(0, 99_999_999):08d}"
        if ncm != "00000000": return ncm

def normalizar_quantidade(qty_str: str, allow_stock: bool) -> Optional[int]:
    try: qty = int(float(qty_str))
    except (ValueError, TypeError): qty = None
    return qty if allow_stock else None

def analisar_dimensoes(bruto: str) -> List[Optional[float]]:
    partes = [p.strip().replace(',', '.') for p in bruto.split('x')]
    def to_float(s):
        try: return float(s)
        except: return None
    return [to_float(p) for p in partes]

def preencher_dimensoes(ws, row, cabecalhos: Dict[str,int], bruto_dim_str: str, campos_valor: List[str], campos_unidade: Optional[List[str]] = None, unidade_padrao: str = "centímetros"):
    partes = [p.strip().replace(',', '.') for p in bruto_dim_str.split('x')]
    valores = []
    tem_valores_validos = False
    
    for p in partes:
        try:
            valor = float(p)
            valores.append(valor)
            if valor > 0: 
                tem_valores_validos = True
        except ValueError:
            valores.append(None)
    
    _celula = ws.cell
    
    # Preencher os valores das dimensões
    for indice, cabecalho in enumerate(campos_valor):
        col = cabecalhos.get(cabecalho)
        if not col:
            continue
        valor = valores[indice] if indice < len(valores) else None
        if valor is not None:
            _celula(row=row, column=col, value=valor)
    
    # CORREÇÃO: Preencher unidades APENAS para campos específicos de pacote e item
    if tem_valores_validos and campos_unidade:
        for indice, cabecalho_unidade in enumerate(campos_unidade):
            if indice < len(valores) and valores[indice] is not None:
                # Só preencher unidades para campos específicos de pacote e item
                if any(termo in cabecalho_unidade for termo in ["pacote", "item"]):
                    col = cabecalhos.get(cabecalho_unidade)
                    if col:
                        _celula(row=row, column=col, value=unidade_padrao)

def tem_atributos_variacao(product):
    """Verifica se o produto possui atributos de variação"""
    atributos_variacao = ['cor', 'tamanho', 'size', 'color', 'modelo', 'estilo']
    
    for atributo in atributos_variacao:
        if product.get(atributo) and str(product.get(atributo)).strip().lower() not in ['', 'nan', 'none']:
            return True
    
    # Verifica se há múltiplos produtos com mesmo SKU base (indicando variações)
    sku = product.get('sku', '')
    if '-' in sku:  # SKUs de variação geralmente têm formato SKU-VARIACAO
        return True
        
    return False

def determinar_tema_variacao(product):
    """Determina o tema de variação baseado nos atributos do produto"""
    tem_cor = bool(product.get('cor') or product.get('color'))
    tem_tamanho = bool(product.get('tamanho') or product.get('size'))
    
    if tem_cor and tem_tamanho:
        return 'SizeColor'
    elif tem_tamanho:
        return 'SizeName'
    elif tem_cor:
        return 'ColorName'
    else:
        return None

def extrair_sku_pai(sku):
    """Extrai o SKU pai removendo sufixos de variação"""
    if '-' in sku:
        return sku.split('-')[0]
    return sku

def preencher_campos_peso(ws, cabecalhos, bruto_peso_value, unit_map, row):
    def normalizar_numero_interno(bruto):
        try: return float(str(bruto).replace(',', '.'))
        except (ValueError, TypeError): return None
    peso_val = normalizar_numero_interno(bruto_peso_value)
    if peso_val is None: return
    for cabecalho_nome, col_indice in cabecalhos.items():
        if "Peso" in cabecalho_nome and (cabecalho_nome.endswith("peso") or cabecalho_nome.startswith("item_package_weight") or cabecalho_nome.endswith("item")):
            ws.cell(row=row, column=col_indice, value=peso_val)
    # Correção: NÃO preencher unidades automaticamente
    # As unidades devem ser preenchidas apenas quando explicitamente necessário

def normalizar_numero(bruto: str) -> float:
    try: return float(bruto.replace(',', '.'))
    except (ValueError, TypeError): return 0.0

def normalizar_unidade(cabecalho: str, bruto_unit: str) -> str: return MAPA_UNIDADES.get(cabecalho.strip(), "").lower().capitalize()

def processar_string_produto_pai(texto_original: Union[str, float]) -> str:
    texto_str = str(texto_original); texto_sem_produto_pai = texto_str.replace("Produto pai:", ""); partes = texto_sem_produto_pai.split("Variações:"); texto_final = partes[0]
    return texto_final.strip()

def set_cell_value(ws, row, header_map, header_name, value):
    if not value or str(value).strip() == '': return
    col_index = header_map.get(header_name)
    if col_index: ws.cell(row=row, column=col_index, value=value)
    else: pass

# ===== FUNÇÕES OTIMIZADAS PARA PROCESSAMENTO EM LOTE =====

def batch_process_main_content(products_data: list, batch_size: int = 5, force_regenerate: bool = False) -> dict:
    """
    Processa múltiplos produtos em lote para geração de conteúdo principal.
    Integrado com sistema de memória inteligente para reutilização de conteúdo.
    
    Args:
        products_data: Lista de dados dos produtos
        batch_size: Tamanho do lote para processamento
        force_regenerate: Se True, força regeneração mesmo se existir na memória
    
    Returns:
        Dicionário mapeando índice do produto para o conteúdo gerado
    """
    print(f"INFO: Iniciando processamento em lote de {len(products_data)} produtos (lotes de {batch_size})")
    print(f"INFO: Modo regeneração forçada: {'Ativado' if force_regenerate else 'Desativado'}")
    
    # Verificar quais produtos já existem na memória
    if not force_regenerate:
        print("INFO: Verificando produtos existentes na memória...")
        cached_products = batch_check_products_in_memory(products_data)
        print(f"INFO: Encontrados {len(cached_products)} produtos na memória")
    else:
        cached_products = {}
    
    # Separar produtos que precisam ser processados
    products_to_process = []
    results_map = {}
    
    for i, product_data in enumerate(products_data):
        identifier = extract_product_identifier(product_data)
        
        if not force_regenerate and identifier in cached_products:
            # Usar conteúdo da memória
            cached_data = cached_products[identifier]
            generated_content = cached_data.get('generated_content', {})
            results_map[i] = generated_content
            print(f"INFO: Produto {i} ({identifier}) reutilizado da memória")
        else:
            # Adicionar à lista para processamento
            products_to_process.append((i, product_data))
    
    if not products_to_process:
        print("INFO: Todos os produtos foram encontrados na memória. Nenhum processamento necessário.")
        return results_map
    
    print(f"INFO: {len(products_to_process)} produtos precisam ser processados")
    
    # Preparar requisições para produtos que precisam ser processados
    requests = []
    for i, product_data in products_to_process:
        product_context = format_product_context(product_data)
        prompt = f"Gerar conteúdo principal para: {product_context}"
        
        requests.append({
            'prompt': prompt,
            'context': {
                'product_index': i,
                'product_title': product_data.get('titulo', f'Produto {i}'),
                'product_data': product_data
            }
        })
    
    # Processar em lotes usando cache inteligente
    def batch_ai_function(batch_requests):
        results = []
        main_ia_chain = get_main_ia_chain()
        
        for request in batch_requests:
            try:
                product_context = format_product_context(request['context']['product_data'])
                parsed_content = main_ia_chain.invoke({"product_context": product_context})
                data_to_return = parsed_content.dict() if hasattr(parsed_content, 'dict') else parsed_content
                results.append(data_to_return)
            except Exception as e:
                print(f"Erro ao processar produto {request['context']['product_index']}: {e}")
                results.append({})
        
        return results
    
    # Executar processamento em lote apenas para produtos necessários
    if requests:
        batch_results = batch_ai_requests(requests, batch_ai_function, batch_size)
        
        # Organizar resultados e salvar na memória
        for idx, (original_index, product_data) in enumerate(products_to_process):
            if idx < len(batch_results):
                generated_content = batch_results[idx] if batch_results[idx] is not None else {}
                results_map[original_index] = generated_content
                
                # Salvar na memória para reutilização futura
                try:
                    save_generated_content_to_memory(
                        product_data=product_data,
                        generated_content=generated_content,
                        force_update=force_regenerate
                    )
                except Exception as e:
                    print(f"AVISO: Erro ao salvar produto na memória: {e}")
            else:
                results_map[original_index] = {}
    
    print(f"INFO: Processamento em lote concluído. {len(results_map)} produtos processados.")
    print(f"INFO: {len([k for k in results_map.keys() if k in [i for i, _ in products_to_process]])} novos produtos gerados")
    print(f"INFO: {len(cached_products)} produtos reutilizados da memória")
    return results_map

def batch_process_field_choices(products_data: list, temp_file_path: str, batch_size: int = 3) -> dict:
    """
    Processa múltiplos produtos em lote para escolha de opções de campos.
    
    Args:
        products_data: Lista de dados dos produtos
        temp_file_path: Caminho do arquivo temporário da planilha
        batch_size: Tamanho do lote para processamento
    
    Returns:
        Dicionário mapeando índice do produto para as escolhas de campos
    """
    print(f"INFO: Iniciando processamento em lote de escolhas de campos para {len(products_data)} produtos")
    
    wb = None
    try:
        wb = load_workbook(filename=temp_file_path, keep_vba=True)
        ws = wb["Modelo"]
        field_options = coletar_opcoes_campo(wb, "Modelo", 7)
        multi_value_fields = identificar_campos_multi_valor(ws)
        vectorstore = get_vectorstore()
        retriever = vectorstore.as_retriever()
        
        # Preparar campos para IA
        fields_for_ai_batch = [
            {"field_name": name, "options": opts, "multi_value": name in multi_value_fields, "is_critical": name in campos_criticos}
            for name, campo_obj in field_options.items()
            if (opts := [o for o in campo_obj['options'] if o and o.lower() != 'nan'])
        ]
        
        if not fields_for_ai_batch:
            print("INFO: Nenhum campo de seleção para processar em lote.")
            return {i: {} for i in range(len(products_data))}
        
        # Preparar requisições para processamento em lote
        requests = []
        for i, product_data in enumerate(products_data):
            titulo_produto = product_data.get('titulo', f'Produto {i}')
            
            requests.append({
                'prompt': f"Escolher opções para produto: {titulo_produto}",
                'context': {
                    'product_index': i,
                    'product_data': product_data,
                    'fields_for_ai_batch': fields_for_ai_batch
                }
            })
        
        # Processar em lotes
        def batch_choices_ai_function(batch_requests):
            results = []
            for request in batch_requests:
                try:
                    product_data = request['context']['product_data']
                    ai_choices, _ = escolher_com_ia(product_data, fields_for_ai_batch, frozenset(), retriever)
                    results.append(ai_choices)
                except Exception as e:
                    print(f"Erro ao processar escolhas para produto {request['context']['product_index']}: {e}")
                    results.append({})
            return results
        
        batch_results = batch_ai_requests(requests, batch_choices_ai_function, batch_size)
        
        # Organizar resultados
        results_map = {}
        for i, result in enumerate(batch_results):
            results_map[i] = result if result is not None else {}
        
        print(f"INFO: Processamento em lote de escolhas concluído. {len(results_map)} produtos processados.")
        return results_map
        
    except Exception as e:
        print(f"ERRO no processamento em lote de escolhas: {e}")
        return {i: {} for i in range(len(products_data))}
    finally:
        if wb:
            wb.close()

def batch_process_chunks(products_data: list, temp_file_path: str, campos_criticos_list: list, persona: str, batch_size: int = 3) -> dict:
    """
    Processa múltiplos produtos em lote para preenchimento de chunks.
    
    Args:
        products_data: Lista de dados dos produtos
        temp_file_path: Caminho do arquivo temporário da planilha
        campos_criticos_list: Lista de campos críticos
        persona: Persona do especialista
        batch_size: Tamanho do lote para processamento
    
    Returns:
        Dicionário mapeando (product_index, chunk_name) para os dados preenchidos
    """
    print(f"INFO: Iniciando processamento em lote de chunks para {len(products_data)} produtos")
    
    wb = None
    try:
        wb = load_workbook(filename=temp_file_path, keep_vba=True)
        ws = wb["Modelo"]
        chunks = mapear_chunks_da_planilha(ws)
        vectorstore = get_vectorstore()
        retriever = vectorstore.as_retriever()
        campos_criticos = set(campos_criticos_list)
        
        results_map = {}
        
        # Processar cada chunk separadamente em lotes
        for chunk_name, chunk_data in chunks.items():
            print(f"INFO: Processando chunk '{chunk_name}' em lote...")
            
            # Preparar requisições para este chunk
            requests = []
            for i, product_data in enumerate(products_data):
                titulo_produto = product_data.get('titulo', f'Produto {i}')
                product_context_str = format_product_context(product_data)
                retriever_context_info = "Contexto da base de conhecimento..."
                
                requests.append({
                    'prompt': f"Processar chunk {chunk_name} para produto: {titulo_produto}",
                    'context': {
                        'product_index': i,
                        'chunk_name': chunk_name,
                        'product_data': product_data,
                        'chunk_data': chunk_data.to_dict(),
                        'retriever_context_info': retriever_context_info,
                        'campos_criticos': campos_criticos,
                        'persona': persona
                    }
                })
            
            # Processar em lotes
            def batch_chunk_ai_function(batch_requests):
                results = []
                for request in batch_requests:
                    try:
                        ctx = request['context']
                        result = processar_chunk_com_ia(
                            ctx['chunk_name'],
                            ctx['chunk_data'],
                            ctx['product_data'],
                            ctx['retriever_context_info'],
                            ctx['campos_criticos'],
                            ctx['persona']
                        )
                        results.append(result)
                    except Exception as e:
                        print(f"Erro ao processar chunk {ctx['chunk_name']} para produto {ctx['product_index']}: {e}")
                        results.append({})
                return results
            
            batch_results = batch_ai_requests(requests, batch_chunk_ai_function, batch_size)
            
            # Armazenar resultados
            for i, result in enumerate(batch_results):
                key = (i, chunk_name)
                results_map[key] = result if result is not None else {}
        
        print(f"INFO: Processamento em lote de chunks concluído. {len(results_map)} resultados gerados.")
        return results_map
        
    except Exception as e:
        print(f"ERRO no processamento em lote de chunks: {e}")
        return {}
    finally:
        if wb:
            wb.close()

def preencher_dados_fixos(ws, row, product, image_urls, cabecalho4, cabecalho5):
    print(f"INFO: Aplicando dados do formulário para o produto SKU {product.get('sku')}")
    set_cell_value(ws, row, cabecalho4, "SKU", product.get("sku"))
    nome_modelo = product.get("nome_marca") if product.get("tipo_marca") == "Marca" else product.get("titulo")
    set_cell_value(ws, row, cabecalho4, "Nome do Modelo", nome_modelo)
    set_cell_value(ws, row, cabecalho4, "Preço sugerido com impostos", product.get("preco"))
    marca_valor = "Genérico" if product.get("tipo_marca") in ("Genérico", "") else product.get("nome_marca")
    set_cell_value(ws, row, cabecalho4, "Fabricante", marca_valor)
    set_cell_value(ws, row, cabecalho4, "Nome da marca", marca_valor)
    
    # Correção: produtos genéricos NÃO devem ter ID do produto preenchido
    if product.get("tipo_marca") != "Genérico":
        set_cell_value(ws, row, cabecalho4, "ID do produto", product.get("id_produto"))
        # CORREÇÃO: NÃO preencher automaticamente "Tipo de ID do produto"
        # Este campo deve ser preenchido apenas quando explicitamente necessário
    
    allow_stock = False
    if product.get("fba_dba", "").upper() == "FBA":
        set_cell_value(ws, row, cabecalho4, "Código do canal de processamento (BR)", "AMAZON_NA")
        allow_stock = True
    else:
        set_cell_value(ws, row, cabecalho4, "Código do canal de processamento (BR)", "DEFAULT")
    
    qty = normalizar_quantidade(product.get("quantidade"), allow_stock)
    set_cell_value(ws, row, cabecalho4, "Quantidade (BR)", qty)
    ncm_val = normalizar_ncm(product.get("ncm")) or gerar_ncm_aleatorio()
    set_cell_value(ws, row, cabecalho4, "Código NCM", ncm_val)
    
    # Correção: NÃO preencher automaticamente campos de hierarquia e variação
    # Estes campos devem ser preenchidos apenas quando explicitamente necessário
    # Removido preenchimento automático de:
    # - Nível de hierarquia
    # - Tipo de relação com o secundário  
    # - Nome do tema de variação
    
    # CORREÇÃO: Preencher dimensões do pacote (sem unidades automáticas)
    if product.get("c_l_a_pacote"):
        preencher_dimensoes(ws, row, cabecalho4, product["c_l_a_pacote"], 
                            ["Comprimento do pacote", "Largura do pacote", "Altura do pacote"],
                            ["Unidade de comprimento do pacote", "Unidade de largura do pacote", "Unidade de altura do pacote"],
                            "centímetros")
    
    # CORREÇÃO: peso do pacote obrigatório com valor padrão e unidade
    peso_pacote = product.get("peso_pacote")
    if not peso_pacote:
        peso_pacote = "100"  # Valor padrão em gramas
        print(f"AVISO: Peso do pacote não informado para SKU {product.get('sku')}, usando valor padrão: 100g")
    
    peso_val = None
    try:
        peso_val = float(str(peso_pacote).replace(',', '.'))
    except (ValueError, TypeError):
        peso_val = 100.0  # Valor padrão
    
    if peso_val:
        set_cell_value(ws, row, cabecalho4, "Peso do pacote", peso_val)
        # CORREÇÃO: Preencher unidade de peso do pacote
        set_cell_value(ws, row, cabecalho4, "Unidade de peso do pacote", "gramas")
    
    # CORREÇÃO: Preencher dimensões do item (sem unidades automáticas)
    if product.get("c_l_a_produto"):
        preencher_dimensoes(ws, row, cabecalho4, product["c_l_a_produto"], 
                            ["Comprimento do item", "Largura do item", "Altura do item"],
                            ["Unidade de comprimento do item", "Unidade de largura do item", "Unidade de altura do item"],
                            "centímetros")
    
    # CORREÇÃO: Preencher País de origem (valor padrão: Brasil)
    set_cell_value(ws, row, cabecalho4, "País de origem", "Brasil")
    
    if image_urls.get('principal'):
        set_cell_value(ws, row, cabecalho4, "URL da imagem principal", image_urls['principal'])
    if image_urls.get('amostra'):
        set_cell_value(ws, row, cabecalho4, "URL da imagem de amostra", image_urls['amostra'])
    
    extra_cols = [cell.column for cell in ws[5] if cell.value and str(cell.value).startswith("other_product_image_locator")]
    for j, url in enumerate(image_urls.get('extra', [])):
        if j < len(extra_cols):
            ws.cell(row=row, column=extra_cols[j], value=url)
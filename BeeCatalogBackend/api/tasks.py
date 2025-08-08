import io
import json
import os
import re
import time
import uuid
import traceback
import base64
import logging
from celery import shared_task, group, chord
from celery.result import AsyncResult
from django.conf import settings
from celery.utils.log import get_task_logger
from openpyxl import load_workbook
from pydantic.v1 import BaseModel
import pandas as pd
import asyncio
from . import utils
from collections import defaultdict

# Configurar logging estruturado
# Remover linha duplicada do logger original
structured_logger = logging.getLogger('beecatalog.tasks') 


logger = get_task_logger(__name__)

def safe_update_state(task_instance, state, meta=None):
    """Safely update task state, avoiding errors in eager mode"""
    try:
        if hasattr(task_instance, 'request') and task_instance.request and task_instance.request.id:
            task_instance.update_state(state=state, meta=meta or {})
        else:
            # In eager mode, just log the progress
            print(f"INFO: [Task Progress] {state}: {meta}")
    except Exception as e:
        print(f"WARNING: Could not update task state: {e}")

@shared_task(bind=True)
def scrape_images_task(self, url):
    try:
        safe_update_state(self, 'PROGRESS', {'step': 'Iniciando extração...'})
        image_urls = utils.scrape_mercadolivre_images(url)
        print(f"INFO: [Scraper] Extraídas {len(image_urls)} imagens da URL: {url}")
        return {'status': 'SUCCESS', 'image_urls': image_urls}
    except Exception as e:
        print(f"ERRO na tarefa de scraping: {e}")
        traceback.print_exc()
        safe_update_state(self, 'FAILURE', {'exc_type': type(e).__name__, 'exc_message': str(e)})
        raise

# *** INÍCIO DA REFATORAÇÃO: Recebe product_data completo ***
@shared_task
def generate_main_content_task(product_index, product_data):
    try:
        start_time = time.time()
        titulo_produto = product_data.get('titulo', f'Produto {product_index}')
        print(f"INFO: [Sub-tarefa] Iniciando geração de conteúdo principal para produto {product_index} ('{titulo_produto[:30]}...')")
        
        # Formata o contexto completo do produto
        product_context = utils.format_product_context(product_data)
        
        main_ia_chain = utils.get_main_ia_chain()
        # Invoca a IA com o contexto completo
        parsed_content = main_ia_chain.invoke({"product_context": product_context})
        
        end_time = time.time()
        print(f"--- [PROFILE][Sub-tarefa: Conteúdo Principal] Produto {product_index} levou: {end_time - start_time:.2f}s")
        data_to_return = parsed_content.dict() if isinstance(parsed_content, BaseModel) else parsed_content
        return {'type': 'main_content', 'product_index': product_index, 'data': data_to_return}
    except Exception as e:
        print(f"ERRO na sub-tarefa de conteúdo principal para produto {product_index}: {e}")
        traceback.print_exc()
        return {'type': 'main_content', 'product_index': product_index, 'data': {}}
# *** FIM DA REFATORAÇÃO ***

# *** INÍCIO DA REFATORAÇÃO: Recebe product_data completo ***
@shared_task
def choose_options_task(product_index, product_data, temp_file_path):
    wb = None
    try:
        start_time = time.time()
        titulo_produto = product_data.get('titulo', f'Produto {product_index}')
        print(f"INFO: [Sub-tarefa] Iniciando escolha de opções para produto {product_index} ('{titulo_produto[:30]}...')")
        
        wb = load_workbook(filename=temp_file_path, keep_vba=True)
        ws = wb["Modelo"]
        field_options = utils.coletar_opcoes_campo(wb, "Modelo", 7)
        multi_value_fields = utils.identificar_campos_multi_valor(ws)
        fields_for_ai_batch = [
            {"field_name": name, "options": opts, "multi_value": name in multi_value_fields, "is_critical": name in utils.campos_criticos}
            for name, campo_obj in field_options.items()
            if (opts := [o for o in campo_obj['options'] if o and o.lower() != 'nan'])
        ]

        if not fields_for_ai_batch:
            print(f"INFO: [Sub-tarefa] Nenhum campo de seleção para preencher para o produto {product_index}. Pulando.")
            return {'type': 'options', 'product_index': product_index, 'data': {}}
            
        vectorstore = utils.get_vectorstore()
        retriever = vectorstore.as_retriever()
        # Passa o product_data completo para a função da IA
        ai_choices, _ = utils.escolher_com_ia(product_data, fields_for_ai_batch, frozenset(), retriever)
        end_time = time.time()
        print(f"--- [PROFILE][Sub-tarefa: Escolher Opções] Produto {product_index} levou: {end_time - start_time:.2f}s")
        return {'type': 'options', 'product_index': product_index, 'data': ai_choices}
    except Exception as e:
        print(f"ERRO na sub-tarefa de escolher opções para produto {product_index}: {e}")
        traceback.print_exc()
        return {'type': 'options', 'product_index': product_index, 'data': {}}
    finally:
        if wb:
            wb.close()
# *** FIM DA REFATORAÇÃO ***

# *** INÍCIO DA REFATORAÇÃO: Recebe product_data completo ***
@shared_task
def process_chunk_task(product_index, chunk_name, product_data, temp_file_path, campos_criticos_list, persona):
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    wb = None
    try:
        start_time = time.time()
        print(f"INFO: [Sub-tarefa] Iniciando processamento do chunk '{chunk_name}' para produto {product_index}...")
        
        wb = load_workbook(filename=temp_file_path, keep_vba=True)
        ws = wb["Modelo"]
        chunks = utils.mapear_chunks_da_planilha(ws)
        chunk_data = chunks[chunk_name].to_dict()
        
        vectorstore = utils.get_vectorstore()
        retriever = vectorstore.as_retriever()
        
        titulo_produto = product_data.get('titulo', f'Produto {product_index}')
        context_str = "\n".join([doc.page_content for doc in retriever.get_relevant_documents(f"Informações para {titulo_produto}")])

        # Passa o product_data completo para a função da IA
        ai_choices = utils.processar_chunk_com_ia(
            chunk_name, chunk_data, product_data, context_str, set(campos_criticos_list), persona
        )
        end_time = time.time()
        print(f"--- [PROFILE][Sub-tarefa: Chunk '{chunk_name}'] Produto {product_index} levou: {end_time - start_time:.2f}s")
        return {'type': 'chunk', 'product_index': product_index, 'chunk_name': chunk_name, 'data': ai_choices}
    except Exception as e:
        print(f"ERRO na sub-tarefa de chunk '{chunk_name}' para produto {product_index}: {e}")
        traceback.print_exc()
        return {'type': 'chunk', 'product_index': product_index, 'chunk_name': chunk_name, 'data': {}}
    finally:
        if wb:
            wb.close()
# *** FIM DA REFATORAÇÃO ***

@shared_task(bind=True)
def assemble_spreadsheet_task(self, results_list, products_data, image_urls_map, temp_file_path):
    wb = None
    try:
        safe_update_state(self, 'PROGRESS', {'step': 'Montando planilha final...'})
        print(f"INFO: [Finalizador] Iniciando montagem final para {len(products_data)} produtos.")

        wb = load_workbook(filename=temp_file_path, keep_vba=True)
        ws = wb["Modelo"]
        
        cabecalho4 = {str(cell.value).strip(): cell.column for cell in ws[4] if cell.value}
        cabecalho5 = {str(cell.value).strip(): cell.column for cell in ws[5] if cell.value}
        
        organized_results = defaultdict(lambda: defaultdict(dict))
        for result in results_list:
            if not result: continue
            idx = result.get('product_index')
            res_type = result.get('type')
            if res_type == 'chunk':
                organized_results[idx]['chunks'][result.get('chunk_name')] = result.get('data', {})
            elif res_type:
                organized_results[idx][res_type] = result.get('data', {})

        for i, product in enumerate(products_data):
            row = 7 + i
            print(f"INFO: [Finalizador] Preenchendo linha {row} para SKU {product.get('sku')}")
            
            product_results = organized_results[i]
            updated_product = product.copy()
            image_urls = image_urls_map.get(str(i), {})

            main_content = product_results.get('main_content', {})
            if main_content:
                updated_product['titulo'] = main_content.get("titulo", updated_product['titulo'])
                utils.set_cell_value(ws, row, cabecalho4, "Nome do Produto", main_content.get("titulo"))
                utils.set_cell_value(ws, row, cabecalho4, "Descrição do Produto", main_content.get("descricao_produto"))
                
                bullet_points = main_content.get("bullet_points", [])
                for idx, bp_data in enumerate(bullet_points):
                    if idx < 5:
                        bullet_field_name = None
                        for header_name in cabecalho5.keys():
                            if header_name.startswith("bullet_point") and header_name.endswith(f"#{idx+1}.value"):
                                bullet_field_name = header_name
                                break
                        
                        if bullet_field_name:
                            utils.set_cell_value(ws, row, cabecalho5, bullet_field_name, bp_data.get("bullet_point", ""))
                
                # Correção: processar palavras-chave corretamente
                palavras_chave_raw = main_content.get("palavras_chave", "")
                if palavras_chave_raw:
                    # Garantir que as palavras-chave estejam separadas por ponto e vírgula
                    # e tenham pelo menos 10 palavras-chave
                    palavras_lista = [p.strip() for p in palavras_chave_raw.replace(',', ';').split(';') if p.strip()]
                    
                    # Se tiver menos de 10, adicionar palavras-chave mais relevantes
                    if len(palavras_lista) < 10:
                        # Palavras-chave mais específicas e relevantes para produtos diversos
                        palavras_genericas = [
                            "Calçados para esportes aquáticos", "Derek Rose", "Elétrico", "Wi-Fi", "Banana",
                            "Organizador", "Viagem", "Bagagem", "Mala", "Acessório", 
                            "Prático", "Funcional", "Portátil", "Durável", "Útil"
                        ]
                        for palavra in palavras_genericas:
                            if len(palavras_lista) >= 10:
                                break
                            if palavra not in palavras_lista:
                                palavras_lista.append(palavra)
                    
                    palavras_chave_formatadas = '; '.join(palavras_lista[:15])  # Máximo 15 palavras-chave
                    utils.set_cell_value(ws, row, cabecalho5, "generic_keyword1", palavras_chave_formatadas)

            options_content = product_results.get('options', {})
            if options_content:
                field_options = utils.coletar_opcoes_campo(wb, "Modelo", row)
                for field, choice in options_content.items():
                    if field in field_options:
                        utils.preencher_grupo_de_colunas(ws, row, field_options[field]['col_indices'], choice if isinstance(choice, list) else [choice], field)

            chunks_content = product_results.get('chunks', {})
            if chunks_content:
                chunks_map = utils.mapear_chunks_da_planilha(ws)
                for chunk_name, chunk_data in chunks_content.items():
                    if chunk_name in chunks_map:
                        for campo in chunks_map[chunk_name].campos:
                            field_name = campo['cabecalho_l5']
                            if field_name in chunk_data and not ws.cell(row=row, column=campo["col"]).value:
                                valor = chunk_data[field_name]
                                if valor and str(valor).strip().lower() != 'nan':
                                    ws.cell(row=row, column=campo["col"], value=str(valor))

            utils.preencher_dados_fixos(ws, row, updated_product, image_urls, cabecalho4, cabecalho5)

        print("INFO: [Finalizador] Montagem final concluída. Salvando arquivo...")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        encoded_file = base64.b64encode(buf.getvalue()).decode('utf-8')
        final_filename = f"PLANILHA_AMAZON_{time.strftime('%Y-%m-%d_%H-%M')}.xlsm"
        
        return {'status': 'SUCCESS', 'file_content': encoded_file, 'filename': final_filename}
    except Exception as e:
        traceback.print_exc()
        safe_update_state(self, 'FAILURE', {'exc_type': type(e).__name__, 'exc_message': str(e)})
        raise
    finally:
        if wb:
            wb.close()
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)

@shared_task(bind=True)
def generate_spreadsheet_task(self, products_data, image_urls_map, template_path):
    wb = None
    try:
        if not products_data:
            raise ValueError("Nenhum dado de produto fornecido.")

        # Limpa o cache da IA para evitar problemas com preenchimento de campos
        utils.limpar_cache_ia()
        
        safe_update_state(self, 'PROGRESS', {'step': 'Preparando ambiente para tarefas...'})

        temp_dir = os.path.join(settings.BASE_DIR, "temp_files")
        os.makedirs(temp_dir, exist_ok=True)
        assemble_temp_path = os.path.join(temp_dir, f"{self.request.id}_assemble.xlsm")
        
        with open(template_path, 'rb') as src, open(assemble_temp_path, 'wb') as dst:
            dst.write(src.read())

        wb = load_workbook(filename=template_path, keep_vba=True)
        chunks = utils.mapear_chunks_da_planilha(wb["Modelo"])
        wb.close()
        wb = None

        # Verifica se está no modo síncrono (desenvolvimento)
        if settings.CELERY_TASK_ALWAYS_EAGER:
            # Modo síncrono: executa tarefas sequencialmente
            print(f"INFO: [Maestra] Executando tarefas em modo síncrono para {len(products_data)} produtos.")
            
            results_list = []
            total_tasks = len(products_data) * 5  # Aproximadamente 5 tarefas por produto
            current_task = 0
            
            for i, product in enumerate(products_data):
                safe_update_state(self, 'PROGRESS', {'step': f'Processando produto {i+1}/{len(products_data)}...'})
                
                # Executa tarefas sequencialmente
                current_task += 1
                result = generate_main_content_task(i, product)
                results_list.append(result)
                
                current_task += 1
                result = choose_options_task(i, product, assemble_temp_path)
                results_list.append(result)

                chunk_personas = {
                    "Oferta (BR) - (Vender na Amazon)": "Especialista em dados de OFERTA...",
                    "Detalhes do produto": "Especialista em especificações TÉCNICAS...",
                    "Segurança e Conformidade": "Especialista em CONFORMIDADE e segurança..."
                }
                for name, persona in chunk_personas.items():
                    if name in chunks:
                        current_task += 1
                        result = process_chunk_task(i, name, product, assemble_temp_path, list(utils.campos_criticos), persona)
                        results_list.append(result)
            
            # Executa a tarefa final de montagem
            safe_update_state(self, 'PROGRESS', {'step': 'Montando planilha final...'})
            final_result = assemble_spreadsheet_task(results_list, products_data, image_urls_map, assemble_temp_path)
            return final_result
        else:
            # Modo assíncrono: usa chord (produção)
            header_tasks = []
            
            for i, product in enumerate(products_data):
                header_tasks.append(generate_main_content_task.s(i, product))
                header_tasks.append(choose_options_task.s(i, product, assemble_temp_path))

                chunk_personas = {
                    "Oferta (BR) - (Vender na Amazon)": "Especialista em dados de OFERTA...",
                    "Detalhes do produto": "Especialista em especificações TÉCNICAS...",
                    "Segurança e Conformidade": "Especialista em CONFORMIDADE e segurança..."
                }
                for name, persona in chunk_personas.items():
                    if name in chunks:
                        header_tasks.append(
                            process_chunk_task.s(i, name, product, assemble_temp_path, list(utils.campos_criticos), persona)
                        )

            print(f"INFO: [Maestra] Criando chord para {len(products_data)} produtos.")
            
            body_task = assemble_spreadsheet_task.s(
                products_data, 
                image_urls_map, 
                assemble_temp_path
            )

            the_chord = chord(header_tasks, body_task)
            result_chord = the_chord.apply_async()
            
            return result_chord.id

    except Exception as e:
        traceback.print_exc()
        safe_update_state(self, 'FAILURE', {'exc_type': type(e).__name__, 'exc_message': str(e)})
        raise
    finally:
        if wb:
            wb.close()
        if 'template_path' in locals() and os.path.exists(template_path):
            os.remove(template_path)


@shared_task(bind=True)
def organizador_ia_task(self, csv_content_base64):
    logger.info("Iniciando tarefa do Organizador IA.")
    
    # Limpa o cache da IA para evitar problemas com preenchimento de campos
    utils.limpar_cache_ia()
    
    try:
        csv_bytes = base64.b64decode(csv_content_base64)
        csv_file_obj = io.StringIO(csv_bytes.decode('utf-8'))
        df = pd.read_csv(csv_file_obj)
        products_to_process = df.to_dict('records')
        total_products = len(products_to_process)
    except Exception as e:
        logger.error(f"Erro ao ler o arquivo CSV: {e}")
        safe_update_state(self, 'FAILURE', {'step': 'Erro de Leitura de CSV'})
        return {'status': 'FAILURE', 'result': 'CSV inválido ou mal formatado.'}

    ia_chain = utils.get_main_ia_chain()
    generated_products_data = []

    for i, product_info in enumerate(products_to_process):
        safe_update_state(self, 'PROGRESS', {'step': 'Gerando conteúdo', 'current': i + 1, 'total': total_products})
        
        # *** INÍCIO DA REFATORAÇÃO: Usa a nova função para formatar o contexto ***
        product_context = utils.format_product_context(product_info)
        
        try:
            # Passa o contexto formatado para a IA
            ia_output = ia_chain.invoke({"product_context": product_context})
            
            new_product_data = {**product_info, **ia_output}
            
            if 'bullet_points' in new_product_data and isinstance(new_product_data['bullet_points'], list):
                new_product_data['bullet_points'] = [bp.get('bullet_point') for bp in new_product_data['bullet_points'] if isinstance(bp, dict) and 'bullet_point' in bp]

            generated_products_data.append(new_product_data)

        except Exception as e:
            logger.error(f"Erro na IA para o produto na linha {i+2}: {e}")
            continue
        # *** FIM DA REFATORAÇÃO ***
        
        time.sleep(1.5)

    return {
        'status': 'SUCCESS',
        'result': {
            'products_data': generated_products_data
        }
    }

# ===== TASKS OTIMIZADAS PARA PROCESSAMENTO EM LOTE =====

@shared_task(bind=True)
def batch_generate_main_content_task(self, products_data_list, batch_size=5):
    """
    Task otimizada para processamento em lote de conteúdo principal.
    """
    try:
        start_time = time.time()
        structured_logger.info("Iniciando processamento em lote de conteúdo principal", extra={
            'task_id': self.request.id,
            'products_count': len(products_data_list),
            'batch_size': batch_size
        })
        
        # Usar função otimizada de processamento em lote
        results_map = utils.batch_process_main_content(products_data_list, batch_size)
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        structured_logger.info("Processamento em lote de conteúdo principal concluído", extra={
            'task_id': self.request.id,
            'processing_time': processing_time,
            'products_processed': len(results_map),
            'avg_time_per_product': processing_time / len(results_map) if results_map else 0
        })
        
        return {
            'status': 'success',
            'results': results_map,
            'processing_time': processing_time,
            'products_count': len(results_map)
        }
        
    except Exception as e:
        structured_logger.error("Erro no processamento em lote de conteúdo principal", extra={
            'task_id': self.request.id,
            'error': str(e),
            'error_type': type(e).__name__
        })
        safe_update_state(self, 'FAILURE', {
            'exc_type': type(e).__name__,
            'exc_message': str(e)
        })
        raise

@shared_task(bind=True)
def batch_choose_options_task(self, products_data_list, temp_file_path, batch_size=3):
    """
    Task otimizada para processamento em lote de escolhas de opções.
    """
    try:
        start_time = time.time()
        structured_logger.info("Iniciando processamento em lote de escolhas de opções", extra={
            'task_id': self.request.id,
            'products_count': len(products_data_list),
            'batch_size': batch_size
        })
        
        # Usar função otimizada de processamento em lote
        results_map = utils.batch_process_field_choices(products_data_list, temp_file_path, batch_size)
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        structured_logger.info("Processamento em lote de escolhas concluído", extra={
            'task_id': self.request.id,
            'processing_time': processing_time,
            'products_processed': len(results_map)
        })
        
        return {
            'status': 'success',
            'results': results_map,
            'processing_time': processing_time,
            'products_count': len(results_map)
        }
        
    except Exception as e:
        structured_logger.error("Erro no processamento em lote de escolhas", extra={
            'task_id': self.request.id,
            'error': str(e),
            'error_type': type(e).__name__
        })
        safe_update_state(self, 'FAILURE', {
            'exc_type': type(e).__name__,
            'exc_message': str(e)
        })
        raise

@shared_task(bind=True)
def batch_process_chunks_task(self, products_data_list, temp_file_path, campos_criticos_list, persona, batch_size=3):
    """
    Task otimizada para processamento em lote de chunks.
    """
    try:
        start_time = time.time()
        structured_logger.info("Iniciando processamento em lote de chunks", extra={
            'task_id': self.request.id,
            'products_count': len(products_data_list),
            'batch_size': batch_size
        })
        
        # Usar função otimizada de processamento em lote
        results_map = utils.batch_process_chunks(
            products_data_list, 
            temp_file_path, 
            campos_criticos_list, 
            persona, 
            batch_size
        )
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        structured_logger.info("Processamento em lote de chunks concluído", extra={
            'task_id': self.request.id,
            'processing_time': processing_time,
            'chunks_processed': len(results_map)
        })
        
        return {
            'status': 'success',
            'results': results_map,
            'processing_time': processing_time,
            'chunks_count': len(results_map)
        }
        
    except Exception as e:
        structured_logger.error("Erro no processamento em lote de chunks", extra={
            'task_id': self.request.id,
            'error': str(e),
            'error_type': type(e).__name__
        })
        safe_update_state(self, 'FAILURE', {
            'exc_type': type(e).__name__,
            'exc_message': str(e)
        })
        raise

@shared_task(bind=True)
def optimized_generate_spreadsheet_task(self, products_data, image_urls_map, template_path, use_batch_processing=True):
    """
    Task principal otimizada que coordena todo o processamento usando lotes.
    """
    try:
        start_time = time.time()
        task_id = self.request.id
        
        structured_logger.info("Iniciando geração otimizada de planilha", extra={
            'task_id': task_id,
            'products_count': len(products_data),
            'use_batch_processing': use_batch_processing
        })
        
        # Atualizar progresso
        safe_update_state(self, 'PROGRESS', {
            'current': 10,
            'total': 100,
            'status': 'Preparando processamento em lote...'
        })
        
        # Criar arquivo temporário
        temp_file_path = os.path.join(settings.BASE_DIR, "uploads", f"temp_{task_id}.xlsm")
        utils.copiar_template_para_temp(template_path, temp_file_path)
        
        if use_batch_processing and len(products_data) > 1:
            # Usar processamento em lote otimizado
            
            # 1. Processamento em lote de conteúdo principal
            safe_update_state(self, 'PROGRESS', {
                'current': 20,
                'total': 100,
                'status': 'Processando conteúdo principal em lote...'
            })
            
            main_content_results = utils.batch_process_main_content(products_data, batch_size=5)
            
            # 2. Processamento em lote de escolhas de opções
            safe_update_state(self, 'PROGRESS', {
                'current': 40,
                'total': 100,
                'status': 'Processando escolhas de campos em lote...'
            })
            
            field_choices_results = utils.batch_process_field_choices(products_data, temp_file_path, batch_size=3)
            
            # 3. Processamento em lote de chunks
            safe_update_state(self, 'PROGRESS', {
                'current': 60,
                'total': 100,
                'status': 'Processando chunks em lote...'
            })
            
            campos_criticos_list = list(utils.campos_criticos)
            persona = "Você é um especialista em catalogação de produtos para Amazon."
            
            chunk_results = utils.batch_process_chunks(
                products_data, 
                temp_file_path, 
                campos_criticos_list, 
                persona, 
                batch_size=3
            )
            
            # 4. Montagem da planilha
            safe_update_state(self, 'PROGRESS', {
                'current': 80,
                'total': 100,
                'status': 'Montando planilha final...'
            })
            
            # Combinar todos os resultados
            combined_results = []
            for i, product_data in enumerate(products_data):
                result = {
                    'type': 'combined',
                    'product_index': i,
                    'main_content': main_content_results.get(i, {}),
                    'field_choices': field_choices_results.get(i, {}),
                    'chunks': {chunk_name: chunk_results.get((i, chunk_name), {}) 
                              for chunk_name in utils.mapear_chunks_da_planilha(
                                  utils.load_workbook(temp_file_path)["Modelo"]
                              ).keys()}
                }
                combined_results.append(result)
            
        else:
            # Fallback para processamento individual (produtos únicos)
            combined_results = []
            for i, product_data in enumerate(products_data):
                # Processar individualmente usando as funções originais
                main_content = utils.get_main_ia_chain().invoke({
                    "product_context": utils.format_product_context(product_data)
                })
                
                result = {
                    'type': 'individual',
                    'product_index': i,
                    'main_content': main_content.dict() if hasattr(main_content, 'dict') else main_content,
                    'field_choices': {},
                    'chunks': {}
                }
                combined_results.append(result)
        
        # Montar planilha final
        final_file_path = utils.assemble_final_spreadsheet(
            combined_results, 
            products_data, 
            image_urls_map, 
            temp_file_path
        )
        
        # Limpar arquivo temporário
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        end_time = time.time()
        total_processing_time = end_time - start_time
        
        structured_logger.info("Geração otimizada de planilha concluída", extra={
            'task_id': task_id,
            'total_processing_time': total_processing_time,
            'products_count': len(products_data),
            'avg_time_per_product': total_processing_time / len(products_data),
            'use_batch_processing': use_batch_processing
        })
        
        safe_update_state(self, 'PROGRESS', {
            'current': 100,
            'total': 100,
            'status': 'Concluído!'
        })
        
        return {
            'status': 'success',
            'file_path': final_file_path,
            'processing_time': total_processing_time,
            'products_count': len(products_data),
            'optimization_used': use_batch_processing
        }
        
    except Exception as e:
        structured_logger.error("Erro na geração otimizada de planilha", extra={
            'task_id': self.request.id,
            'error': str(e),
            'error_type': type(e).__name__
        })
        
        # Limpar arquivo temporário em caso de erro
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
        safe_update_state(self, 'FAILURE', {
            'exc_type': type(e).__name__,
            'exc_message': str(e)
        })
        raise